"""Script de importación inicial: productos y leads desde el Excel de Corze."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from database import FirebaseDB
from datetime import datetime

EXCEL = 'datos_iniciales/inventario/Matriz de Gestion CORZE_22.xlsx'

def fmt_date(v):
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    return str(v)[:10] if v else ''

def run():
    db = FirebaseDB()
    wb = openpyxl.load_workbook(EXCEL, data_only=True)

    creados = 0
    errores = 0

    # ── PANELES ──────────────────────────────────────────────────────────────
    ws = wb['Inventario Paneles']
    for row in ws.iter_rows(min_row=4, values_only=True):
        sku = row[0]
        if not sku or not str(sku).startswith('PAN'): continue
        ok, _ = db.add_producto({
            'codigo': str(sku),
            'nombre': str(row[16] or f"{row[1]} {row[2]}"),
            'categoria': 'Panel Solar',
            'descripcion': f"Tecnología: {row[4]} | Eficiencia: {int((row[5] or 0)*100)}% | Dimensiones: {row[6]} | Peso: {row[7]}kg",
            'marca': str(row[1] or ''),
            'modelo': str(row[2] or ''),
            'potencia_w': int(row[3] or 0),
            'precio_costo': 0,
            'precio_venta': int(row[10] or 0),
            'unidad': 'unidad',
            'stock_actual': int(row[12] or 0),
            'stock_minimo': int(row[13] or 0),
            'activo': True,
            'notas': f"Garantía producto: {row[8]} años | Garantía rendimiento: {row[9]} años | Proveedor: {row[11]}",
        })
        if ok: creados += 1
        else: errores += 1

    print(f"Paneles: {creados} creados, {errores} errores")
    c_prev = creados; creados = 0; errores = 0

    # ── INVERSORES ───────────────────────────────────────────────────────────
    ws = wb['Inventario Inversores']
    for row in ws.iter_rows(min_row=4, values_only=True):
        sku = row[0]
        if not sku or not str(sku).startswith('INV'): continue
        ok, _ = db.add_producto({
            'codigo': str(sku),
            'nombre': str(row[14] or f"{row[1]} {row[2]}"),
            'categoria': 'Inversor',
            'descripcion': f"Tipo: {row[3]} | Fase: {row[4]} | Eficiencia: {int((row[6] or 0)*100)}%",
            'marca': str(row[1] or ''),
            'modelo': str(row[2] or ''),
            'potencia_w': int((row[5] or 0) * 1000),
            'precio_costo': 0,
            'precio_venta': int(row[8] or 0),
            'unidad': 'unidad',
            'stock_actual': int(row[10] or 0),
            'stock_minimo': int(row[11] or 0),
            'activo': True,
            'notas': f"Garantía: {row[7]} años | Proveedor: {row[9]}",
        })
        if ok: creados += 1
        else: errores += 1

    print(f"Inversores: {creados} creados, {errores} errores")
    creados = 0; errores = 0

    # ── BATERÍAS ─────────────────────────────────────────────────────────────
    ws = wb['Inventario Baterias']
    for row in ws.iter_rows(min_row=4, values_only=True):
        sku = row[0]
        if not sku or not str(sku).startswith('BAT'): continue
        ok, _ = db.add_producto({
            'codigo': str(sku),
            'nombre': str(row[14] or f"{row[1]} {row[2]}"),
            'categoria': 'Batería',
            'descripcion': f"Química: {row[3]} | Capacidad: {row[4]}kWh | Voltaje: {row[5]}V | Ciclos: {row[6]}",
            'marca': str(row[1] or ''),
            'modelo': str(row[2] or ''),
            'potencia_w': int((row[4] or 0) * 1000),
            'precio_costo': 0,
            'precio_venta': int(row[8] or 0),
            'unidad': 'unidad',
            'stock_actual': int(row[10] or 0),
            'stock_minimo': int(row[11] or 0),
            'activo': True,
            'notas': f"Garantía: {row[7]} años | Proveedor: {row[9]}",
        })
        if ok: creados += 1
        else: errores += 1

    print(f"Baterías: {creados} creados, {errores} errores")
    creados = 0; errores = 0

    # ── ESTRUCTURAS ──────────────────────────────────────────────────────────
    ws = wb['Estructuras']
    for row in ws.iter_rows(min_row=4, values_only=True):
        sku = row[0]
        if not sku or not str(sku).startswith('EST'): continue
        ok, _ = db.add_producto({
            'codigo': str(sku),
            'nombre': str(row[10] or f"{row[1]} - {row[2]}"),
            'categoria': 'Estructura/Montaje',
            'descripcion': f"Material: {row[2]} | Compatibilidad: hasta {row[3]} paneles",
            'marca': '',
            'modelo': str(row[1] or ''),
            'potencia_w': 0,
            'precio_costo': 0,
            'precio_venta': int(row[4] or 0),
            'unidad': 'unidad',
            'stock_actual': int(row[6] or 0),
            'stock_minimo': int(row[7] or 0),
            'activo': True,
            'notas': f"Proveedor: {row[5]}",
        })
        if ok: creados += 1
        else: errores += 1

    print(f"Estructuras: {creados} creados, {errores} errores")
    creados = 0; errores = 0

    # ── MATERIALES / BOS ─────────────────────────────────────────────────────
    ws = wb['Materiales']
    for row in ws.iter_rows(min_row=4, values_only=True):
        sku = row[0]
        if not sku or not str(sku).startswith('MAT'): continue
        ok, _ = db.add_producto({
            'codigo': str(sku),
            'nombre': str(row[2] or ''),
            'categoria': f"Material - {row[1]}",
            'descripcion': str(row[1] or ''),
            'marca': '',
            'modelo': str(row[2] or ''),
            'potencia_w': 0,
            'precio_costo': 0,
            'precio_venta': int(row[4] or 0),
            'unidad': str(row[3] or 'unidad'),
            'stock_actual': int(row[6] or 0),
            'stock_minimo': int(row[7] or 0),
            'activo': True,
            'notas': f"Proveedor: {row[5]}",
        })
        if ok: creados += 1
        else: errores += 1

    print(f"Materiales: {creados} creados, {errores} errores")
    creados = 0; errores = 0

    # ── TRABAJADORES ─────────────────────────────────────────────────────────
    trabajadores = [
        {'nombre': 'Alexis Iván', 'apellido': 'Zamorano Toledo',  'cargo': 'Instalador', 'color': '#3498DB'},
        {'nombre': 'Daniel',      'apellido': 'Zelada Muñoz',      'cargo': 'Instalador', 'color': '#2ECC71'},
        {'nombre': 'Matías Ignacio', 'apellido': 'Olivares Farías', 'cargo': 'Instalador', 'color': '#9B59B6'},
    ]
    for t in trabajadores:
        initials = t['nombre'][0].upper() + t['apellido'][0].upper()
        ok, _ = db.add_trabajador({**t, 'email': '', 'telefono': '', 'rut': '',
                                    'activo': True, 'avatar': initials,
                                    'fecha_ingreso': '', 'notas': 'Importado desde contrato'})
        if ok: creados += 1
        else: errores += 1

    print(f"Trabajadores: {creados} creados, {errores} errores")
    creados = 0; errores = 0

    # ── LEADS EXISTENTES ─────────────────────────────────────────────────────
    ws = wb['Clientes y Cotizaciones']
    etapa_map = {
        'Vendido': 'Proyecto Finalizado',
        'En diseño': 'Propuesta Enviada',
        'Cotizado': 'Propuesta Enviada',
        'En instalación': 'En Instalación',
        'Terminado': 'Proyecto Finalizado',
        'Prospecto': 'Nuevo Lead',
    }
    for row in ws.iter_rows(min_row=4, values_only=True):
        cot = row[0]
        if not cot or not str(cot).startswith('COT') or not row[3]: continue
        estado_raw = str(row[13] or '')
        etapa = etapa_map.get(estado_raw, 'Nuevo Lead')
        ok, _ = db.add_lead({
            'nombre': str(row[3] or '').split()[0],
            'apellido': ' '.join(str(row[3] or '').split()[1:]),
            'email': str(row[7] or ''),
            'telefono': str(row[6] or ''),
            'rut': str(row[4] or ''),
            'empresa': '',
            'comuna': str(row[11] or ''),
            'region': str(row[12] or ''),
            'origen': 'excel',
            'etapa': etapa,
            'asignado_a': str(row[14] or ''),
            'tipo_proyecto': str(row[9] or 'Residencial'),
            'consumo_kwh': 0,
            'notas': f"Cotización original: {cot} | Proyecto: {row[8]} | Dirección: {row[10]}",
            'presupuestos_ids': [],
            'historial': [],
        })
        if ok: creados += 1
        else: errores += 1

    print(f"Leads: {creados} creados, {errores} errores")
    print("\n✓ Importación completada")

if __name__ == '__main__':
    run()
