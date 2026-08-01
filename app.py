# app.py — VTA Web v2.1

from flask import (Flask, render_template, request, redirect,
                   url_for, session, jsonify, flash, Response, make_response)
from flask.json.provider import DefaultJSONProvider
from functools import wraps
import os, json, threading, imaplib
import email as _email_lib
from email.header import decode_header as _email_decode_header
from email.utils import parseaddr as _parseaddr
from datetime import datetime, date, timezone, timedelta

_TZ = timezone(timedelta(hours=-4))
def _cl(): return datetime.now(_TZ).replace(tzinfo=None)

from config import USERS, APP_NAME, APP_VERSION
from database import FirebaseDB

class FirestoreJSONProvider(DefaultJSONProvider):
    """Serializa DatetimeWithNanoseconds y otros tipos Firestore a string."""
    def default(self, o):
        if isinstance(o, (datetime, date)):
            return o.isoformat()[:19]
        # Firestore DatetimeWithNanoseconds / Timestamp
        if hasattr(o, 'isoformat'):
            return o.isoformat()[:19]
        if hasattr(o, '_seconds'):      # google.protobuf Timestamp
            return datetime.utcfromtimestamp(o._seconds).strftime('%Y-%m-%d %H:%M:%S')
        try:
            return str(o)
        except:
            return None

app = Flask(__name__)
app.json_provider_class = FirestoreJSONProvider
app.json = FirestoreJSONProvider(app)
app.secret_key = os.environ.get('SECRET_KEY', 'corze-secret-2024')
app.config['SESSION_PERMANENT'] = False
db = FirebaseDB()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

@app.context_processor
def inject_globals():
    import config as cfg
    usuario = session.get('usuario', '')
    role = USERS.get(usuario, {}).get('role', 'admin')
    return {
        'app_name': APP_NAME, 'app_version': APP_VERSION,
        'usuario': usuario,
        'usuario_color': session.get('usuario_color', '#F5A623'),
        'usuario_avatar': session.get('usuario_avatar', ''),
        'usuario_role': role,
        'config': cfg,
    }

# Endpoints permitidos para usuarios con rol "flipping"
_FLIPPING_ALLOWED = {
    'login', 'logout', 'index', 'vitrina', 'api_vitrina', 'privacidad',
    'flipping',
    'api_flipping_list', 'api_flipping_add', 'api_flipping_get',
    'api_flipping_update', 'api_flipping_delete', 'api_flipping_simular',
    'car_hunter',
    'api_carhunter_list', 'api_carhunter_add', 'api_carhunter_get',
    'api_carhunter_update', 'api_carhunter_delete', 'api_carhunter_seed_kr',
    'api_carhunter_clear_kr',
    'service_worker', 'manifest', 'static',
    'api_public_portada', 'api_public_fotos', 'api_contacto_vitrina',
    'webhook_verify', 'webhook_receive',
    'api_validar_rut', 'add_cache_headers',
    'api_buscar_patente', 'api_buscar_patente_poll',
}

@app.before_request
def check_limited_access():
    if 'usuario' not in session:
        return None
    user_data = USERS.get(session.get('usuario', ''), {})
    if user_data.get('role') == 'flipping':
        endpoint = request.endpoint or ''
        if endpoint not in _FLIPPING_ALLOWED:
            if request.path.startswith('/api/'):
                return jsonify({'ok': False, 'error': 'Acceso limitado — no tienes permisos para esta sección'}), 403
            return render_template('acceso_limitado.html'), 403

def _build_vitrina_cars():
    """Lista de autos para la vitrina (una sola lectura Firestore, sin fotos embebidas)."""
    from datetime import timedelta
    try:
        all_docs = db.get_all_inventario(strip_photos=True) or []
    except Exception as e:
        print(f'Error _build_vitrina_cars: {e}')
        return []

    cutoff   = (_cl() - timedelta(days=90)).strftime('%Y-%m-%d')
    publicados = [r for r in all_docs if r.get('estado') == 'Publicado']
    vendidos   = []
    for r in [r for r in all_docs if r.get('estado') == 'Vendido']:
        fv = r.get('fecha_venta') or ''
        if hasattr(fv, 'strftime'):    fv = fv.strftime('%Y-%m-%d')
        elif hasattr(fv, 'isoformat'): fv = fv.isoformat()[:10]
        else: fv = str(fv)[:10]
        if fv >= cutoff:
            vendidos.append(r)

    ahora = _cl()
    cars  = []
    for r in publicados + vendidos:
        try:
            precio = int(r.get('precio_vitrina') or 0)
            if not precio:
                precio = int(r.get('precio_pedido') or 0) \
                    if r.get('tipo_registro') == 'CONSIGNACIÓN' \
                    else int(r.get('precio_venta_colaboradores') or 0)
            badge = ''
            if r.get('estado') == 'Vendido':
                badge = 'vendido'
            elif r.get('nuevo_precio'):
                badge = 'nuevo_precio'
            else:
                try:
                    fi = r.get('fecha_ingreso', '')
                    fi_str = fi.strftime('%Y-%m-%d') if hasattr(fi, 'strftime') else str(fi)[:10]
                    if fi_str and (ahora - datetime.strptime(fi_str, '%Y-%m-%d')).days <= 7:
                        badge = 'recien_llegado'
                except Exception:
                    pass
            cars.append({
                'id':               r.get('id'),
                'marca':            r.get('marca', ''),
                'modelo':           r.get('modelo', ''),
                'anio':             r.get('anio'),
                'tipo_vehiculo':    r.get('tipo_vehiculo', ''),
                'color':            r.get('color', ''),
                'transmision':      r.get('transmision', ''),
                'traccion':         r.get('traccion', ''),
                'combustible':      r.get('combustible', ''),
                'km_aprox':         r.get('km_aprox'),
                'cantidad_duenos':  r.get('cantidad_duenos'),
                'motor':            r.get('motor', ''),
                'patente':          r.get('patente', ''),
                'tipo_registro':    r.get('tipo_registro', ''),
                'estado':           r.get('estado', ''),
                'fecha_ingreso':    str(r.get('fecha_ingreso', ''))[:10],
                'precio_vitrina':   precio,
                'descripcion_publica': r.get('descripcion_publica', ''),
                'total_fotos':      int(r.get('total_fotos', 0)),
                'badge':            badge,
            })
        except Exception as e:
            print(f'Error auto vitrina {r.get("id","?")}: {e}')
    return cars


# ── AUTH ──────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return redirect(url_for('dashboard') if 'usuario' in session else url_for('login'))

@app.route('/admin', methods=['GET'])
def admin():
    return redirect(url_for('dashboard') if 'usuario' in session else url_for('login'))

@app.route('/api/validar-rut', methods=['GET'])
@login_required
def api_validar_rut():
    """Valida RUT chileno y obtiene datos desde múltiples fuentes"""
    import requests as req

    rut_raw = request.args.get('rut', '').strip().upper()
    rut     = rut_raw.replace('.', '').replace('-', '').replace(' ', '')

    if not rut or len(rut) < 2:
        return jsonify({'valid': False, 'error': 'RUT vacío'})

    numero = rut[:-1]
    dv     = rut[-1]

    if not numero.isdigit():
        return jsonify({'valid': False, 'error': 'Formato inválido'})

    # ── Calcular dígito verificador ───────────────────────────────────────
    suma, mul = 0, 2
    for c in reversed(numero):
        suma += int(c) * mul
        mul = 2 if mul == 7 else mul + 1
    res     = 11 - (suma % 11)
    dv_calc = '0' if res == 11 else ('K' if res == 10 else str(res))
    valid   = dv == dv_calc

    # Formatear
    fmt = f"{int(numero):,}".replace(',', '.') + '-' + dv_calc

    result = {
        'valid':          valid,
        'rut_formateado': fmt,
        'dv_esperado':    dv_calc,
        'numero':         numero,
        'nombre':         '',
        'tipo':           '',          # persona / empresa
        'actividad_sii':  '',
        'inicio_actividades': '',
        'giro':           '',
        'estado_sii':     '',
        'fuente':         '',
    }

    if not valid:
        return jsonify(result)

    # Determinar si es empresa o persona (empresas < 50.000.000)
    es_empresa = int(numero) > 50_000_000 or int(numero) < 1_000_000

    # ── 1. Consultar LibreAPI (nombre) ────────────────────────────────────
    try:
        r = req.get(f'https://api.libreapi.cl/rut/activities?rut={numero}-{dv_calc}',
                    timeout=4, headers={'Accept': 'application/json'})
        if r.status_code == 200:
            data = r.json()
            result['nombre']             = data.get('name', '')
            result['actividad_sii']      = data.get('activity', '')
            result['inicio_actividades'] = data.get('start_activities_date', '')
            result['giro']               = data.get('category', '')
            result['estado_sii']         = data.get('status', '')
            result['tipo']               = 'empresa' if data.get('category') else 'persona'
            result['fuente']             = 'SII via LibreAPI'
    except Exception:
        pass

    # ── 2. Fallback: rutify.cl ────────────────────────────────────────────
    if not result['nombre']:
        try:
            r = req.get(f'https://api.rutify.cl/rut/{numero}',
                        timeout=4, headers={'Accept': 'application/json'})
            if r.status_code == 200:
                data = r.json()
                result['nombre'] = data.get('razon_social') or data.get('nombre') or ''
                result['fuente'] = 'rutify.cl'
        except Exception:
            pass

    # ── 3. Determinar tipo por número si no se pudo ───────────────────────
    if not result['tipo']:
        num_int = int(numero)
        result['tipo'] = 'empresa' if num_int >= 50_000_000 or num_int < 900_000 else 'persona'

    return jsonify(result)

@app.route('/privacidad')
def privacidad():
    return render_template('privacidad.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data    = request.get_json(silent=True) or {}
        usuario = data.get('usuario', '').strip()
        pin     = data.get('pin', '').strip()
        if usuario not in USERS:
            return jsonify({'ok': False, 'msg': 'Usuario no encontrado'})
        if pin != USERS[usuario]['pin']:
            return jsonify({'ok': False, 'msg': '⛔  PIN incorrecto — inténtalo de nuevo'})
        session['usuario']         = usuario
        session['usuario_color']   = USERS[usuario]['color']
        session['usuario_avatar']  = USERS[usuario]['avatar']
        return jsonify({'ok': True, 'msg': f'✅  Acceso autorizado — Bienvenido/a, {usuario}'})
    return render_template('login.html', users=USERS)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ── PÁGINAS ───────────────────────────────────────────────────────────────────
@app.route('/admin/dashboard')
@login_required
def dashboard():
    try:
        stats = db.get_corze_dashboard_stats()
    except Exception as e:
        stats = {}
        flash(str(e), 'error')
    return render_template('dashboard.html', stats=stats, page='dashboard')

@app.route('/admin/inventario')
@login_required
def inventario():
    return render_template('inventario.html', page='inventario')

# Redirigir rutas antiguas
@app.route('/admin/compraventa')
@login_required
def compraventa():
    return redirect(url_for('inventario'))

@app.route('/admin/consignaciones')
@login_required
def consignaciones():
    return redirect(url_for('inventario'))

# ── API INVENTARIO (unificado) ────────────────────────────────────────────────
@app.route('/api/inventario', methods=['GET'])
@login_required
def api_inv_list():
    tipo   = request.args.get('tipo', '')
    estado = request.args.get('estado', '')
    q      = request.args.get('q', '')
    rows   = db.get_all_inventario(tipo=tipo, estado=estado, query=q)
    return jsonify(rows)

@app.route('/api/inventario', methods=['POST'])
@login_required
def api_inv_add():
    data = request.get_json()
    data['usuario_creacion'] = session.get('usuario', '')
    ok, err = db.add_inventario(data)
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/inventario/<doc_id>', methods=['GET'])
@login_required
def api_inv_get(doc_id):
    return jsonify(db.get_inventario_by_id(doc_id) or {})

@app.route('/api/inventario/<doc_id>', methods=['PUT'])
@login_required
def api_inv_update(doc_id):
    data = request.get_json()
    data['usuario_ultima_edicion'] = session.get('usuario', '')
    ok, err = db.update_inventario(doc_id, data)
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/inventario/<doc_id>', methods=['DELETE'])
@login_required
def api_inv_delete(doc_id):
    ok = db.delete_inventario(doc_id, session.get('usuario', ''))
    return jsonify({'ok': ok})

@app.route('/api/inventario/<doc_id>/sell', methods=['POST'])
@login_required
def api_inv_sell(doc_id):
    data = request.get_json()
    ok, err = db.quick_sell_inventario(
        doc_id, int(data.get('precio_venta', 0)),
        data.get('fecha_venta', ''), session.get('usuario', ''))
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/inventario/<doc_id>/duplicate', methods=['POST'])
@login_required
def api_inv_duplicate(doc_id):
    ok, err = db.duplicate_inventario(doc_id, session.get('usuario', ''))
    return jsonify({'ok': ok, 'error': err})

# ── EXPORT ────────────────────────────────────────────────────────────────────
@app.route('/api/export/inventario')
@login_required
def api_export_inv():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return Response("Instala openpyxl: pip install openpyxl", status=500)

    rows = db.get_all_inventario()
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "Corze Inventario"

    headers = [
        'Tipo','Patente','Marca','Año','Tipo Vehículo','Modelo','Motor','Chasis','Color',
        'Transmisión','Combustible','KM','N° Dueños',
        'P. Compra / Pedido','P. Colaboradores / Mínimo','V. Mercado',
        'P. Venta Final','Ganancia','Comisión %','Comisión $',
        'Estado','Días Stock','F. Ingreso','F. Publicación','F. Venta',
        'Propietario','Contacto','Notas','Creado por','Editado por',
    ]
    keys = [
        'tipo_registro','patente','marca','anio','tipo_vehiculo','modelo','motor','chasis','color',
        'transmision','combustible','km_aprox','cantidad_duenos',
        'precio_compra','precio_venta_colaboradores','valor_mercado',
        'precio_venta_final','ganancia','comision_porcentaje','comision_monto',
        'estado','dias_en_stock','fecha_ingreso','fecha_publicacion','fecha_venta',
        'nombre_propietario','contacto_propietario','notas','usuario_creacion','usuario_ultima_edicion',
    ]

    # Estilos
    hdr_fill  = PatternFill("solid", fgColor="003D5C")
    hdr_font  = Font(bold=True, color="F5A623", size=10)
    int_fill  = PatternFill("solid", fgColor="0D2030")
    con_fill  = PatternFill("solid", fgColor="2D2500")
    alt_fill  = PatternFill("solid", fgColor="131920")
    def_fill  = PatternFill("solid", fgColor="0E1318")
    thin      = Side(style='thin', color='1E2D3D')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal='center', vertical='center')

    # Cabecera
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[1].height = 22

    # Filas
    for ri, row in enumerate(rows, 2):
        tipo    = str(row.get('tipo_registro', ''))
        bg_fill = int_fill if tipo == 'INTERNO' else (con_fill if tipo == 'CONSIGNACIÓN' else (alt_fill if ri % 2 == 0 else def_fill))

        for ci, key in enumerate(keys, 1):
            val  = row.get(key, '')
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = bg_fill
            cell.font      = Font(color="EEF2F7", size=9)
            cell.border    = border
            cell.alignment = Alignment(vertical='center')

    # Anchos de columna
    widths = [12,10,14,6,16,26,14,18,14,16,14,10,8,
              16,16,14,14,14,8,12,
              10,8,12,12,12,
              20,16,30,14,14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Autofilter
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"

    # Guardar en memoria y retornar
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    ts = _cl().strftime('%Y%m%d_%H%M')
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment;filename=CORZE_Inventario_{ts}.xlsx'}
    )

@app.route('/api/stats')
@login_required
def api_stats():
    return jsonify(db.get_corze_dashboard_stats())

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)

# ══════════════════════════════════════════════════════════════════════════════
#  CRM — Rutas de interfaz y API
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/crm/enviar_directo', methods=['POST'])
@login_required
def api_crm_enviar_directo():
    """Envía un mensaje WhatsApp directo sin necesitar conv_id."""
    data     = request.get_json()
    telefono = data.get('telefono', '').strip()
    mensaje  = data.get('mensaje', '').strip()
    if not telefono or not mensaje:
        return jsonify({'ok': False, 'error': 'Faltan datos'})
    try:
        _enviar_whatsapp(telefono, mensaje)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/admin/crm')
@login_required
def crm():
    return render_template('crm.html', page='crm')

# ── API Conversaciones ────────────────────────────────────────────────────────
@app.route('/api/crm/conversaciones', methods=['GET'])
@login_required
def api_crm_convs():
    return jsonify(db.get_conversaciones())

@app.route('/api/crm/conversaciones/<conv_id>', methods=['PUT'])
@login_required
def api_crm_conv_update(conv_id):
    data = request.get_json()
    ok   = db.update_conversacion(conv_id, data)
    return jsonify({'ok': ok})

@app.route('/api/crm/conversaciones/<conv_id>/leer', methods=['POST'])
@login_required
def api_crm_leer(conv_id):
    db.marcar_leido(conv_id)
    return jsonify({'ok': True})

@app.route('/api/crm/conversaciones/<conv_id>/mensajes', methods=['GET'])
@login_required
def api_crm_mensajes(conv_id):
    return jsonify(db.get_mensajes(conv_id))

@app.route('/api/crm/conversaciones/<conv_id>/enviar', methods=['POST'])
@login_required
def api_crm_enviar(conv_id):
    data     = request.get_json()
    mensaje  = data.get('mensaje', '').strip()
    if not mensaje:
        return jsonify({'ok': False, 'error': 'Mensaje vacío'})

    # Obtener conversación para saber el canal y sender_id
    convs = db.get_conversaciones()
    conv  = next((c for c in convs if c['id'] == conv_id), None)
    if not conv:
        return jsonify({'ok': False, 'error': 'Conversación no encontrada'})

    usuario = session.get('usuario', '')

    try:
        if conv['canal'] == 'whatsapp':
            _enviar_whatsapp(conv['telefono'] or conv['sender_id'], mensaje)
        elif conv['canal'] == 'messenger':
            _enviar_messenger(conv['sender_id'], mensaje)

        db.add_mensaje(conv_id, mensaje, 'saliente',
                       nombre_remitente=usuario, leido=True)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

# ── Envíos a Meta ──────────────────────────────────────────────────────────────
def _enviar_whatsapp(to: str, mensaje: str):
    import requests as req
    token    = os.environ.get('WHATSAPP_TOKEN', '')
    phone_id = os.environ.get('WHATSAPP_PHONE_ID', '')
    if not token or not phone_id:
        raise Exception('WHATSAPP_TOKEN o WHATSAPP_PHONE_ID no configurados')
    url  = f'https://graph.facebook.com/v19.0/{phone_id}/messages'
    resp = req.post(url,
        headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},
        json={'messaging_product': 'whatsapp', 'to': to,
              'type': 'text', 'text': {'body': mensaje}})
    if resp.status_code not in (200, 201):
        raise Exception(f'WhatsApp error {resp.status_code}: {resp.text}')

def _enviar_messenger(sender_id: str, mensaje: str):
    import requests as req
    token = os.environ.get('MESSENGER_TOKEN', '')
    if not token:
        raise Exception('MESSENGER_TOKEN no configurado')
    url  = 'https://graph.facebook.com/v19.0/me/messages'
    resp = req.post(url,
        params={'access_token': token},
        json={'recipient': {'id': sender_id}, 'message': {'text': mensaje}})
    if resp.status_code not in (200, 201):
        raise Exception(f'Messenger error {resp.status_code}: {resp.text}')

# ══════════════════════════════════════════════════════════════════════════════
#  WEBHOOK — Recibe mensajes de Meta
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/webhook/meta', methods=['GET'])
def webhook_verify():
    verify_token = os.environ.get('WEBHOOK_VERIFY_TOKEN', 'corze2024webhook')
    mode      = request.args.get('hub.mode')
    token     = request.args.get('hub.verify_token')
    challenge = request.args.get('hub.challenge')
    if mode == 'subscribe' and token == verify_token:
        print(f'✅ Webhook verificado correctamente')
        return challenge, 200
    print(f'❌ Webhook verificación fallida: mode={mode} token={token}')
    return 'Forbidden', 403

@app.route('/webhook/test', methods=['GET'])
@login_required
def webhook_test():
    """Diagnóstico: verifica configuración del webhook."""
    wa_token   = os.environ.get('WHATSAPP_TOKEN', '')
    wa_phone   = os.environ.get('WHATSAPP_PHONE_ID', '')
    msg_token  = os.environ.get('MESSENGER_TOKEN', '')
    vfy_token  = os.environ.get('WEBHOOK_VERIFY_TOKEN', '')
    return jsonify({
        'webhook_url':             f"{request.host_url}webhook/meta",
        'WHATSAPP_TOKEN':          '✅ configurado' if wa_token else '❌ falta',
        'WHATSAPP_PHONE_ID':       '✅ configurado' if wa_phone else '❌ falta',
        'MESSENGER_TOKEN':         '✅ configurado' if msg_token else '❌ falta',
        'WEBHOOK_VERIFY_TOKEN':    vfy_token or '❌ falta',
        'conversaciones_en_db':    len(db.get_conversaciones()),
    })

@app.route('/webhook/meta', methods=['POST'])
def webhook_receive():
    """Recibe mensajes entrantes de WhatsApp y Messenger."""
    data = request.get_json(silent=True) or {}

    try:
        for entry in data.get('entry', []):
            # ── WhatsApp ──────────────────────────────────────────────────
            for change in entry.get('changes', []):
                val = change.get('value', {})
                for msg in val.get('messages', []):
                    sender   = msg.get('from', '')
                    contenido= msg.get('text', {}).get('body', '[media]')
                    perfil   = val.get('contacts', [{}])[0]
                    nombre   = perfil.get('profile', {}).get('name', '')
                    conv     = db.get_or_create_conversacion(
                        sender_id=sender, canal='whatsapp',
                        nombre=nombre, telefono=sender)
                    if conv:
                        db.add_mensaje(conv['id'], contenido, 'entrante',
                                       nombre_remitente=nombre or sender)

            # ── Messenger ─────────────────────────────────────────────────
            for messaging in entry.get('messaging', []):
                sender_id = messaging.get('sender', {}).get('id', '')
                msg       = messaging.get('message', {})
                contenido = msg.get('text', '[media]')
                if sender_id and contenido:
                    conv = db.get_or_create_conversacion(
                        sender_id=sender_id, canal='messenger')
                    if conv:
                        db.add_mensaje(conv['id'], contenido, 'entrante')

    except Exception as e:
        print(f'Webhook error: {e}')

    return 'OK', 200

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)

# ══════════════════════════════════════════════════════════════════════════════
#  FOTOS — Upload base64 a Firestore
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/admin/consignacion/terreno')
@login_required
def consignacion_terreno():
    """Formulario de inspección en terreno — optimizado para tablet."""
    from config import MARCAS, TIPOS_VEHICULO, TRANSMISIONES, TRACCIONES, COMBUSTIBLES
    return render_template('consignacion_terreno.html',
        marcas=MARCAS, tipos=TIPOS_VEHICULO,
        transmisiones=TRANSMISIONES, tracciones=TRACCIONES,
        combustibles=COMBUSTIBLES,
        usuario=session.get('usuario',''),
        config=type('C', (), {'USERS': USERS})())

@app.route('/api/consignacion/guardar', methods=['POST'])
@login_required
def api_guardar_consignacion():
    """Guarda la consignación en Firestore e inventario."""
    data = request.get_json()
    try:
        # Crear en inventario como CONSIGNACIÓN
        doc_data = {
            'tipo_registro':         'CONSIGNACIÓN',
            'patente':               data.get('patente','').upper(),
            'chasis':                data.get('chasis',''),
            'marca':                 data.get('marca',''),
            'modelo':                data.get('modelo',''),
            'anio':                  int(data.get('anio', 2020)),
            'color':                 data.get('color',''),
            'motor':                 data.get('motor',''),
            'transmision':           data.get('transmision',''),
            'traccion':              data.get('traccion',''),
            'combustible':           data.get('combustible',''),
            'km_aprox':              int(data.get('km_aprox', 0)),
            'cantidad_duenos':       int(data.get('cantidad_duenos', 1)),
            'nombre_propietario':    data.get('nombre_propietario',''),
            'rut_propietario':       data.get('rut_propietario',''),
            'contacto_propietario':  data.get('telefono_propietario',''),
            'email_propietario':     data.get('email_propietario',''),
            'precio_pedido':         int(data.get('precio_pedido', 0)),
            'precio_publicacion':    int(data.get('precio_publicacion', 0)),
            'precio_minimo':         int(data.get('precio_minimo', 0)),
            'comision_porcentaje':   float(data.get('comision_pct', 5)),
            'estado':                'En Stock',
            'equipamiento':          data.get('equipamiento', {}),
            'inspeccion_tecnica':    data.get('inspeccion_tecnica', {}),
            'notas':                 data.get('notas',''),
            'notas_internas':        data.get('notas_internas',''),
            'permiso_circulacion':   data.get('permiso_circulacion',''),
            'soap':                  data.get('soap',''),
            'revision_tecnica':      data.get('revision_tecnica',''),
            'prendas':               data.get('prendas',''),
            'firma_propietario':     data.get('firma',''),
            'exclusividad':          bool(data.get('exclusividad', False)),
            'modalidad':             data.get('modalidad', 'presencial'),
            'email_propietario':     data.get('email_propietario',''),
            'agente':                session.get('usuario',''),
            'fecha_ingreso':         _cl().strftime('%Y-%m-%d'),
        }
        ok, err = db.add_inventario(doc_data)
        if ok:
            return jsonify({'ok': True, 'id': err})
        return jsonify({'ok': False, 'error': err})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.template_filter('clp')
def clp_filter(n):
    try: return '$' + f'{int(n):,}'.replace(',','.')
    except: return '—'

@app.template_filter('clp_fmt')
def clp_fmt_filter(n):
    try: return f'{int(n):,}'.replace(',','.')
    except: return '—'

def get_logo_b64():
    """Devuelve el logo como base64 para embeber en HTML imprimible."""
    import base64, os
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'assets', 'logo_corze.png')
    try:
        with open(logo_path, 'rb') as f:
            return 'data:image/png;base64,' + base64.b64encode(f.read()).decode()
    except:
        return ''

def render_contrato(car, folio, fecha, **kwargs):
    """Helper para renderizar el contrato con todos los datos."""
    precio_pub     = int(car.get('precio_publicacion') or car.get('precio_pedido',0))
    comision_pct   = float(kwargs.get('comision_pct', car.get('comision_porcentaje',5)))
    precio_pub_ovr = int(kwargs.get('precio_pub', precio_pub))
    comision_monto = round(precio_pub_ovr * comision_pct / 100)
    pago           = precio_pub_ovr - comision_monto
    equip      = car.get('equipamiento', {})
    equip_list = [item for cat in equip.values() for item in (cat if isinstance(cat,list) else [])]
    legal_items = [
        ('Permiso de circulación', car.get('permiso_circulacion','')),
        ('SOAP',                   car.get('soap','')),
        ('Revisión técnica',       car.get('revision_tecnica','')),
        ('Prendas',                car.get('prendas','')),
    ]
    return render_template('consignacion_contrato.html',
        folio=folio, fecha=fecha,
        agente=car.get('agente', session.get('usuario','')),
        logo_b64=get_logo_b64(),
        nombre_propietario=car.get('nombre_propietario',''),
        rut_propietario=car.get('rut_propietario',''),
        telefono_propietario=car.get('contacto_propietario',''),
        patente=car.get('patente',''), chasis=car.get('chasis',''),
        marca=car.get('marca',''), modelo=car.get('modelo',''),
        anio=car.get('anio',''), color=car.get('color',''),
        motor=car.get('motor',''), km_aprox=car.get('km_aprox',0),
        transmision=car.get('transmision',''), combustible=car.get('combustible',''),
        traccion=car.get('traccion',''), cantidad_duenos=car.get('cantidad_duenos',1),
        precio_pedido=car.get('precio_pedido',0),
        precio_publicacion=precio_pub_ovr,
        comision_pct=comision_pct,
        comision_monto=comision_monto,
        pago_propietario=pago,
        legal_items=legal_items,
        equipamiento_list=equip_list,
        inspeccion=car.get('inspeccion_tecnica',{}),
        notas=car.get('notas',''),
        firma=car.get('firma_propietario',''),
        exclusividad=car.get('exclusividad', False),
        modalidad=car.get('modalidad','presencial'))

@app.route('/admin/consignacion/contrato')
@login_required
def consignacion_contrato():
    doc_id = request.args.get('id','')
    car    = db.get_inventario_by_id(doc_id) if doc_id else {}
    if not car:
        return '<h2>Contrato no encontrado</h2>', 404
    import random
    folio = f'CORZE-{_cl().strftime("%Y%m")}-{random.randint(1000,9999)}'
    # Override precio/comision si se pasan como parámetros (regeneración)
    kwargs = {}
    if request.args.get('precio_pub'):
        kwargs['precio_pub'] = int(request.args.get('precio_pub',0))
    if request.args.get('comision_pct'):
        kwargs['comision_pct'] = float(request.args.get('comision_pct',5))
    return render_contrato(car, folio, _cl().strftime('%d/%m/%Y %H:%M'), **kwargs)

@app.route('/api/consignacion/enviar-email', methods=['POST'])
@login_required
def api_enviar_email_contrato():
    """Envía el contrato al email del propietario via Resend."""
    api_key = os.environ.get('RESEND_API_KEY','')
    if not api_key:
        return jsonify({'ok': False, 'error': 'RESEND_API_KEY no configurada en Railway'})
    d        = request.get_json()
    doc_id   = d.get('id','')
    car      = db.get_inventario_by_id(doc_id) if doc_id else {}
    email_to = car.get('email_propietario','') or d.get('email','')
    if not email_to:
        return jsonify({'ok': False, 'error': 'El propietario no tiene email registrado'})
    import random
    folio = f'CORZE-{_cl().strftime("%Y%m")}-{random.randint(1000,9999)}'
    html_contrato = render_contrato(car, folio, _cl().strftime('%d/%m/%Y %H:%M'))
    html_email = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px">
      <div style="background:#F5A623;padding:20px;border-radius:10px 10px 0 0;text-align:center">
        <h1 style="color:#fff;margin:0;font-size:22px">VTA — Vende Tu Auto</h1>
      </div>
      <div style="background:#f9f9f9;padding:24px;border:1px solid #ddd;border-top:none;border-radius:0 0 10px 10px">
        <p style="font-size:15px;color:#333">Estimado/a <strong>{car.get('nombre_propietario','')}</strong>,</p>
        <p style="font-size:14px;color:#555;margin-top:12px;line-height:1.6">
          Le informamos que su contrato de consignación con <strong>Corze</strong> 
          ha sido generado exitosamente para el vehículo:
        </p>
        <div style="background:#fff;border:1px solid #e0e0e0;border-radius:8px;padding:14px;margin:16px 0;text-align:center">
          <div style="font-size:18px;font-weight:bold;color:#F5A623">
            {car.get('marca','')} {car.get('modelo','')} {car.get('anio','')}
          </div>
          <div style="font-size:13px;color:#777;margin-top:4px">Patente: {car.get('patente','')}</div>
        </div>
        <p style="font-size:13px;color:#555;line-height:1.6">
          Adjunto encontrará su contrato firmado. Por favor guárdelo como respaldo.
          Si tiene alguna consulta, contáctenos por WhatsApp o al email de VTA.
        </p>
        <div style="margin-top:20px;text-align:center">
          <a href="https://wa.me/56956086490" 
             style="background:#25D366;color:#fff;padding:12px 24px;border-radius:8px;text-decoration:none;font-weight:bold">
            Consultar por WhatsApp
          </a>
        </div>
        <p style="font-size:11px;color:#aaa;margin-top:20px;text-align:center">
          Corze · Santiago, Chile · corze.cl
        </p>
      </div>
    </div>"""
    try:
        resp = http_requests.post('https://api.resend.com/emails',
            headers={'Authorization': f'Bearer {api_key}', 'Content-Type': 'application/json'},
            json={
                'from':    'Corze <noreply@corze.cl>',
                'to':      [email_to],
                'subject': f'Contrato de Consignación VTA — {car.get("marca","")} {car.get("modelo","")} {car.get("anio","")}',
                'html':    html_email,
            }, timeout=10)
        if resp.status_code in (200, 201):
            return jsonify({'ok': True})
        return jsonify({'ok': False, 'error': resp.text})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@login_required
def migrate_crm():
    """Asigna etapa_crm='Nuevo' a todas las conversaciones que no tienen etapa."""
    try:
        convs   = db.get_conversaciones()
        updated = 0
        for c in convs:
            if not c.get('etapa_crm'):
                db.update_conversacion(c['id'], {'etapa_crm': 'Nuevo'})
                updated += 1
        return jsonify({'ok': True, 'updated': updated, 'total': len(convs)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/admin/migrate/fotos')
@login_required
def migrate_fotos():
    """Un uso: sincroniza foto_portada y total_fotos en todos los autos existentes."""
    try:
        inventario = db.get_all_inventario()
        updated = 0
        for car in inventario:
            cid   = car['id']
            fotos = db.get_fotos(cid)
            if not fotos:
                continue
            portada = next((f for f in fotos if f.get('portada')), fotos[0])
            db.db.collection('inventario').document(cid).update({
                'foto_portada': portada.get('foto', ''),
                'total_fotos':  len(fotos),
            })
            updated += 1
        return jsonify({'ok': True, 'updated': updated})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/public/portada/<doc_id>')
def api_public_portada(doc_id):
    """Foto portada de un vehículo — usada para lazy load en la vitrina."""
    try:
        fotos = db.get_fotos(doc_id)
        portada = next((f for f in fotos if f.get('portada')), fotos[0] if fotos else None)
        if not portada:
            return jsonify({'foto': None, 'object_position': 'center center'})
        return jsonify({
            'foto':            portada.get('foto'),
            'object_position': portada.get('object_position', 'center center'),
        })
    except Exception:
        return jsonify({'foto': None, 'object_position': 'center center'})

@app.route('/api/public/fotos/<doc_id>')
def api_public_fotos(doc_id):
    """API pública de fotos para la vitrina — sin login requerido."""
    try:
        fotos = db.get_fotos(doc_id)
        return jsonify([{
            'foto':             f.get('foto'),
            'object_position':  f.get('object_position', 'center center'),
            'portada':          f.get('portada', False),
        } for f in fotos])
    except Exception:
        return jsonify([])

@app.route('/api/inventario/<doc_id>/fotos', methods=['GET'])
@login_required
def api_get_fotos(doc_id):
    return jsonify(db.get_fotos(doc_id))

@app.route('/api/inventario/<doc_id>/fotos', methods=['POST'])
@login_required
def api_add_foto(doc_id):
    data     = request.get_json()
    foto_b64 = data.get('foto')
    nombre   = data.get('nombre', '')
    if not foto_b64:
        return jsonify({'ok': False, 'error': 'Sin imagen'})
    ok, err = db.add_foto(doc_id, foto_b64, nombre, session.get('usuario', ''))
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/inventario/<doc_id>/fotos/<foto_id>', methods=['DELETE'])
@login_required
def api_del_foto(doc_id, foto_id):
    ok = db.delete_foto(doc_id, foto_id)
    return jsonify({'ok': ok})

@app.route('/api/inventario/<doc_id>/fotos/<foto_id>/portada', methods=['POST'])
@login_required
def api_set_portada(doc_id, foto_id):
    ok = db.set_foto_portada(doc_id, foto_id)
    return jsonify({'ok': ok})

@app.route('/api/inventario/<doc_id>/fotos/<foto_id>', methods=['PUT'])
@login_required
def api_update_foto(doc_id, foto_id):
    data = request.get_json()
    ok   = db.update_foto(doc_id, foto_id, data)
    return jsonify({'ok': ok})

@app.route('/api/inventario/<doc_id>/fotos/reorder', methods=['POST'])
@login_required
def api_reorder_fotos(doc_id):
    data = request.get_json()
    ok   = db.reorder_fotos(doc_id, data.get('order', []))
    return jsonify({'ok': ok})

# ── FINANZAS ──────────────────────────────────────────────────────────────────
@app.route('/admin/finanzas')
@login_required
def finanzas():
    return render_template('finanzas.html', page='finanzas')

@app.route('/api/finanzas', methods=['GET'])
@login_required
def api_finanzas_list():
    tipo = request.args.get('tipo', '')
    mes  = request.args.get('mes', '')
    rows = db.get_all_finanzas(tipo=tipo, mes=mes)
    for r in rows:
        r['tiene_comprobante'] = bool(r.pop('comprobante', None))
    return jsonify(rows)

@app.route('/api/finanzas', methods=['POST'])
@login_required
def api_finanzas_add():
    data = request.get_json()
    data['usuario'] = session.get('usuario', '')
    ok, err = db.add_finanza(data)
    return jsonify({'ok': ok, 'id': err if ok else None, 'error': None if ok else err})

@app.route('/api/finanzas/<doc_id>', methods=['PUT'])
@login_required
def api_finanzas_update(doc_id):
    data = request.get_json()
    ok, err = db.update_finanza(doc_id, data)
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/finanzas/<doc_id>', methods=['DELETE'])
@login_required
def api_finanzas_delete(doc_id):
    ok = db.delete_finanza(doc_id)
    return jsonify({'ok': ok})

@app.route('/api/finanzas/resumen', methods=['GET'])
@login_required
def api_finanzas_resumen():
    mes = request.args.get('mes', '')
    return jsonify(db.get_finanzas_resumen(mes=mes))

@app.route('/api/finanzas/<doc_id>/comprobante', methods=['GET'])
@login_required
def api_finanzas_comprobante(doc_id):
    try:
        doc = db.db.collection('finanzas_corze').document(doc_id).get()
        data = doc.to_dict() or {}
        return jsonify({'comprobante': data.get('comprobante', '')})
    except Exception as e:
        return jsonify({'comprobante': '', 'error': str(e)})

# ── COSTOS FIJOS ──────────────────────────────────────────────────────────────
@app.route('/api/costos-fijos', methods=['GET'])
@login_required
def api_costos_fijos_list():
    return jsonify(db.get_costos_fijos())

@app.route('/api/costos-fijos', methods=['POST'])
@login_required
def api_costos_fijos_add():
    data = request.get_json()
    ok, err = db.add_costo_fijo(data)
    return jsonify({'ok': ok, 'id': err if ok else None, 'error': None if ok else err})

@app.route('/api/costos-fijos/<doc_id>', methods=['PUT'])
@login_required
def api_costos_fijos_update(doc_id):
    data = request.get_json()
    ok, err = db.update_costo_fijo(doc_id, data)
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/costos-fijos/<doc_id>', methods=['DELETE'])
@login_required
def api_costos_fijos_delete(doc_id):
    ok = db.delete_costo_fijo(doc_id)
    return jsonify({'ok': ok})

@app.route('/api/costos-fijos/aplicar', methods=['POST'])
@login_required
def api_costos_fijos_aplicar():
    data = request.get_json()
    mes  = data.get('mes', '')
    if not mes:
        return jsonify({'ok': False, 'error': 'Falta el mes'})
    ok, result = db.aplicar_costos_fijos(mes, usuario=session.get('usuario', ''))
    if ok:
        return jsonify({'ok': True, 'count': result})
    return jsonify({'ok': False, 'error': result})

@app.route('/api/costos-fijos/estado', methods=['GET'])
@login_required
def api_costos_fijos_estado():
    mes = request.args.get('mes', '')
    aplicado = db.costos_fijos_aplicados(mes) if mes else False
    return jsonify({'aplicado': aplicado})

# ── VENTAS CORZE SOLAR ───────────────────────────────────────────────────────
@app.route('/api/ventas-corze', methods=['GET'])
@login_required
def api_ventas_corze_list():
    return jsonify(db.get_all_ventas_corze())

@app.route('/api/ventas-corze', methods=['POST'])
@login_required
def api_ventas_corze_add():
    data = request.get_json() or {}
    ok, result = db.add_venta_corze(data)
    return jsonify({'ok': ok, 'id': result if ok else None, 'error': result if not ok else None})

@app.route('/api/ventas-corze/<doc_id>', methods=['PUT'])
@login_required
def api_ventas_corze_update(doc_id):
    data = request.get_json() or {}
    ok, err = db.update_venta_corze(doc_id, data)
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/ventas-corze/<doc_id>', methods=['DELETE'])
@login_required
def api_ventas_corze_delete(doc_id):
    ok = db.delete_venta_corze(doc_id)
    return jsonify({'ok': ok})

# ── GASTOS CORZE SOLAR ────────────────────────────────────────────────────────
@app.route('/api/gastos-corze', methods=['GET'])
@login_required
def api_gastos_corze_list():
    return jsonify(db.get_all_gastos_corze())

@app.route('/api/gastos-corze', methods=['POST'])
@login_required
def api_gastos_corze_add():
    data = request.get_json() or {}
    ok, result = db.add_gasto_corze(data)
    return jsonify({'ok': ok, 'id': result if ok else None, 'error': result if not ok else None})

@app.route('/api/gastos-corze/<doc_id>', methods=['PUT'])
@login_required
def api_gastos_corze_update(doc_id):
    data = request.get_json() or {}
    ok, err = db.update_gasto_corze(doc_id, data)
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/gastos-corze/<doc_id>', methods=['DELETE'])
@login_required
def api_gastos_corze_delete(doc_id):
    ok = db.delete_gasto_corze(doc_id)
    return jsonify({'ok': ok})

@app.route('/api/admin/seed-ventas-gastos', methods=['POST'])
@login_required
def api_seed_ventas_gastos():
    """Importa datos iniciales del Excel a Firestore."""
    from datetime import datetime as dt

    VENTAS = [
        dict(id_venta='VTA-001', fecha_venta='2026-07-17', num_cotizacion='COT-0321',
             cliente='Mario Lopez', proyecto='PRY-001', vendedor='Enrique Corvalán J',
             monto_venta=7000000, forma_pago='Transferencia',
             estado_instalacion='Pendiente instalación', fecha_entrega='2026-07-31',
             comision_pct=0, comision_monto=0, notas=''),
        dict(id_venta='VTA-002', fecha_venta='2026-07-19', num_cotizacion='COT-0322',
             cliente='Nicolas Pinilla Candia', proyecto='PRY-002', vendedor='Enrique Corvalán H',
             monto_venta=8154920, forma_pago='Tarjeta crédito',
             estado_instalacion='Pendiente instalación', fecha_entrega='2026-07-27',
             comision_pct=0, comision_monto=0, notas=''),
    ]

    GASTOS = [
        dict(id_gasto='GTO-001', fecha='2026-06-23', categoria='Transporte y combustible', descripcion='Combustible y peaje Rapel', monto_neto=47800, iva=0, monto_total=47800, metodo_pago='Tarjeta crédito', responsable='Ignacio Corvalán', empresa='Copec', num_documento='Boleta'),
        dict(id_gasto='GTO-002', fecha='2026-06-24', categoria='Indumentaria', descripcion='Ropa corporativa', monto_neto=457327, iva=86892, monto_total=544219, metodo_pago='Tarjeta crédito', responsable='Enrique Corvalán H', empresa='Texora', num_documento='Boleta'),
        dict(id_gasto='GTO-003', fecha='2026-07-06', categoria='Indumentaria', descripcion='Bordado ropa corporativa', monto_neto=94500, iva=17955, monto_total=112455, metodo_pago='Transferencia', responsable='Enrique Corvalán H', empresa='Roca Estampa', num_documento='230'),
        dict(id_gasto='GTO-004', fecha='2026-07-04', categoria='Transporte y combustible', descripcion='Combustible y peaje Rapel', monto_neto=80000, iva=0, monto_total=80000, metodo_pago='Tarjeta crédito', responsable='Enrique Corvalán J', empresa='Copec', num_documento='Boleta'),
        dict(id_gasto='GTO-005', fecha='2026-07-04', categoria='Marketing', descripcion='Promo 6 meses Integralia', monto_neto=2400000, iva=0, monto_total=2400000, metodo_pago='Tarjeta crédito', responsable='Enrique Corvalán H', empresa='Integralia', num_documento='Factura pendiente'),
        dict(id_gasto='GTO-006', fecha='2026-07-01', categoria='Transporte y combustible', descripcion='Compra camioneta L200', monto_neto=9000000, iva=0, monto_total=9000000, metodo_pago='Transferencia', responsable='Empresa', empresa='', num_documento=''),
        dict(id_gasto='GTO-007', fecha='2026-07-08', categoria='Transporte y combustible', descripcion='Arreglo camioneta L200', monto_neto=250000, iva=0, monto_total=250000, metodo_pago='Transferencia', responsable='Enrique Corvalán H', empresa='Dario Lozano del Rio', num_documento='Particular Directo'),
        dict(id_gasto='GTO-008', fecha='2026-07-10', categoria='Transporte y combustible', descripcion='Arreglo camioneta L200', monto_neto=450000, iva=0, monto_total=450000, metodo_pago='Transferencia', responsable='Enrique Corvalán H', empresa='Dario Lozano del Rio', num_documento='Particular Directo'),
        dict(id_gasto='GTO-009', fecha='2026-07-20', categoria='Materiales proyectos', descripcion='Inv Solis 8 / 2 Bat Dyness 16kwh', monto_neto=4462122, iva=847803, monto_total=5309925, metodo_pago='Tarjeta crédito', responsable='Enrique Corvalán H', empresa='Natura Energy', num_documento='15934'),
        dict(id_gasto='GTO-010', fecha='2026-07-21', categoria='Materiales proyectos', descripcion='Materiales Casa 116 y Lopez', monto_neto=241861, iva=45954, monto_total=287815, metodo_pago='Tarjeta crédito', responsable='Enrique Corvalán H', empresa='Ferrelectric', num_documento='197585'),
        dict(id_gasto='GTO-011', fecha='2026-07-21', categoria='Herramientas', descripcion='Rotomartillo, Cascos de seguridad y Broca', monto_neto=268770, iva=51066, monto_total=319836, metodo_pago='Tarjeta crédito', responsable='Enrique Corvalán H', empresa='Pernos Kim', num_documento='768830'),
        dict(id_gasto='GTO-012', fecha='2026-07-21', categoria='Indumentaria', descripcion='Guantes', monto_neto=39960, iva=7592, monto_total=39960, metodo_pago='Tarjeta crédito', responsable='Enrique Corvalán H', empresa='Mercado Libre', num_documento='Boleta'),
        dict(id_gasto='GTO-013', fecha='2026-07-21', categoria='Herramientas', descripcion='Chicharra y dados allen 6mm', monto_neto=25126, iva=4774, monto_total=29900, metodo_pago='Tarjeta crédito', responsable='Enrique Corvalán H', empresa='Pernos Patricio', num_documento='206504'),
        dict(id_gasto='GTO-014', fecha='2026-07-21', categoria='Materiales proyectos', descripcion='Materiales Casa 116 y Lopez', monto_neto=361977, iva=68776, monto_total=430753, metodo_pago='Tarjeta crédito', responsable='Enrique Corvalán H', empresa='Guzman', num_documento=''),
        dict(id_gasto='GTO-015', fecha='2026-07-21', categoria='Materiales proyectos', descripcion='Materiales Casa 116 y Lopez', monto_neto=1752161, iva=332911, monto_total=2085072, metodo_pago='Tarjeta crédito', responsable='Enrique Corvalán H', empresa='Emat', num_documento='Factura pendiente'),
        dict(id_gasto='GTO-016', fecha='2026-07-23', categoria='Materiales proyectos', descripcion='Paneles Lopez', monto_neto=931980, iva=177076, monto_total=1109056, metodo_pago='Tarjeta crédito', responsable='Enrique Corvalán H', empresa='Dark Energy', num_documento='Factura pendiente'),
        dict(id_gasto='GTO-017', fecha='2026-07-23', categoria='Indumentaria', descripcion='Polar y pantalones', monto_neto=70178, iva=13334, monto_total=83512, metodo_pago='Tarjeta crédito', responsable='Enrique Corvalán H', empresa='Texora', num_documento='714374'),
        dict(id_gasto='GTO-018', fecha='2026-07-21', categoria='Materiales proyectos', descripcion='Materiales Casa 116 y Lopez', monto_neto=75068, iva=14263, monto_total=89331, metodo_pago='Tarjeta crédito', responsable='Enrique Corvalán H', empresa='Vitel', num_documento='1497204'),
        dict(id_gasto='GTO-019', fecha='2026-07-23', categoria='Herramientas', descripcion='Knock Out', monto_neto=57800, iva=10982, monto_total=68782, metodo_pago='Tarjeta crédito', responsable='Enrique Corvalán H', empresa='Inse', num_documento='32915'),
        dict(id_gasto='GTO-020', fecha='2026-07-24', categoria='Transporte y combustible', descripcion='Arreglo camioneta L200', monto_neto=100226, iva=0, monto_total=100226, metodo_pago='Transferencia', responsable='Enrique Corvalán H', empresa='Dario Lozano del Rio', num_documento='Particular Directo'),
        dict(id_gasto='GTO-021', fecha='2026-07-23', categoria='Transporte y combustible', descripcion='Flete EMAT - Casa 116', monto_neto=50000, iva=0, monto_total=50000, metodo_pago='Transferencia', responsable='Enrique Corvalán H', empresa='Manuel Bravo (Niño Feliz)', num_documento='Particular Directo'),
    ]

    now = _cl().isoformat()

    def seed_col(col_name, rows, id_field):
        col = db.db.collection(col_name)
        existing = {d.to_dict().get(id_field): d.id for d in col.stream()}
        for row in rows:
            row.setdefault('created_at', now)
            row['updated_at'] = now
            key = row.get(id_field)
            if key and key in existing:
                col.document(existing[key]).set(row)
            else:
                col.add(row)

    seed_col('ventas_corze', VENTAS, 'id_venta')
    seed_col('gastos_corze', GASTOS, 'id_gasto')
    return jsonify({'ok': True, 'ventas': len(VENTAS), 'gastos': len(GASTOS)})

@app.route('/api/admin/seed-accesorios', methods=['POST'])
@login_required
def api_seed_accesorios():
    """Importa accesorios desde lista de materiales a productos_solar."""
    # Precios unitarios = total_lista / cantidad
    ACCESORIOS = [
        dict(codigo='ACC-001', nombre='Riel 5000mm (Versionpro)',                                          categoria='Accesorio', precio_venta=20028,  precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-002', nombre='Midclamp 30mm',                                                     categoria='Accesorio', precio_venta=493,    precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-003', nombre='Endclamp',                                                          categoria='Accesorio', precio_venta=493,    precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-004', nombre='Groundclip 30x50 mm2',                                             categoria='Accesorio', precio_venta=115,    precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-005', nombre='Tornillo de Conexión a Tierra Groundlug',                          categoria='Accesorio', precio_venta=1200,   precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-006', nombre='Perno Partido con Espiga 6AWG',                                    categoria='Accesorio', precio_venta=67,     precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-007', nombre='Tubo EMT 32MMX3MTS Espesor 1,0 MM',                               categoria='Accesorio', precio_venta=2352,   precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-008', nombre='Caja Metal A-11 100X100X65MM 20-25MM Pre Galvanizada c/Tapa',      categoria='Accesorio', precio_venta=1109,   precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-009', nombre='Caja Metal A-01 100X65X65MM 1/2-3/4" Pre Galvanizada c/Tapa',     categoria='Accesorio', precio_venta=907,    precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-010', nombre='Prensa Estopa PG-7 3-6,5MM',                                      categoria='Accesorio', precio_venta=101,    precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-011', nombre='Prensa Estopa PG-16 10-14MM',                                     categoria='Accesorio', precio_venta=1350,   precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-012', nombre='Conector EMT a Flexible 32MM',                                    categoria='Accesorio', precio_venta=1344,   precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-013', nombre='Conector Recto para Flexible con PVC 32MM',                       categoria='Accesorio', precio_venta=979,    precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-014', nombre='Abrazadera Tipo Caddy 25MM para EMT',                             categoria='Accesorio', precio_venta=153,    precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-015', nombre='Base Ajustable Delantera y Trasera 15/30°',                       categoria='Accesorio', precio_venta=7300,   precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-016', nombre='Cable RV-K Negro 6MM Extraflexible',                              categoria='Accesorio', precio_venta=1260,   precio_costo=0, unidad='metro',  stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-017', nombre='Cable Solar FV (H1Z2Z2-K) 4MM',                                  categoria='Accesorio', precio_venta=500,    precio_costo=0, unidad='metro',  stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-018', nombre='Tablero Metálico 400X300X200MM IP66 IK10 RAL 7035',               categoria='Accesorio', precio_venta=43800,  precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-019', nombre='Automático NXB-63H 2X40A 10KA C',                                 categoria='Accesorio', precio_venta=7458,   precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-020', nombre='Repartidor Bipolar 125A (2X25MM + 5X6MM)',                        categoria='Accesorio', precio_venta=3480,   precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
        dict(codigo='ACC-021', nombre='Barra de Conexión DIN 8 Polos Verde Aislada',                    categoria='Accesorio', precio_venta=1186,   precio_costo=0, unidad='unidad', stock_actual=0, stock_minimo=0, potencia_w=0, activo=True),
    ]
    now = _cl().strftime('%Y-%m-%d %H:%M:%S')
    col = db.db.collection('productos_solar')
    existing = {d.to_dict().get('codigo'): d.id for d in col.stream() if d.to_dict().get('codigo','').startswith('ACC')}
    created = updated = 0
    for p in ACCESORIOS:
        p['created_at'] = p.get('created_at', now)
        p['updated_at'] = now
        if p['codigo'] in existing:
            col.document(existing[p['codigo']]).set(p)
            updated += 1
        else:
            col.add(p)
            created += 1
    return jsonify({'ok': True, 'creados': created, 'actualizados': updated, 'total': len(ACCESORIOS)})

# ══════════════════════════════════════════════════════════════════════════════
#  VITRINA PÚBLICA — corze.cl
# ══════════════════════════════════════════════════════════════════════════════
@app.after_request
def add_cache_headers(response):
    """Evita que Cloudflare cachee respuestas dinámicas."""
    if request.path.startswith('/api/') or request.path.startswith('/admin/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma']  = 'no-cache'
        response.headers['Expires'] = '0'
    return response

@app.route('/vitrina')
def vitrina():
    """Página pública de inventario."""
    return render_template('vitrina.html', cars=_build_vitrina_cars())

@app.route('/api/vitrina')
def api_vitrina():
    """API pública — Publicados + Vendidos recientes."""
    from datetime import datetime, timedelta
    try:
        publicados = db.get_all_inventario(estado='Publicado') or []
    except Exception as e:
        print(f'Error vitrina publicados: {e}')
        publicados = []

    try:
        vendidos_all = db.get_all_inventario(estado='Vendido') or []
        cutoff = (_cl() - timedelta(days=90)).strftime('%Y-%m-%d')
        vendidos = []
        for r in vendidos_all:
            fv = r.get('fecha_venta') or ''
            # Normalizar fecha_venta a string
            if hasattr(fv, 'strftime'):
                fv = fv.strftime('%Y-%m-%d')
            elif hasattr(fv, 'isoformat'):
                fv = fv.isoformat()[:10]
            else:
                fv = str(fv)[:10]
            if fv >= cutoff:
                vendidos.append(r)
    except Exception as e:
        print(f'Error vitrina vendidos: {e}')
        vendidos = []

    ahora  = _cl()
    public = []

    for r in publicados + vendidos:
        try:
            precio_vitrina = int(r.get('precio_vitrina') or 0)
            if not precio_vitrina:
                precio_vitrina = int(r.get('precio_pedido') or 0) \
                    if r.get('tipo_registro') == 'CONSIGNACIÓN' \
                    else int(r.get('precio_venta_colaboradores') or 0)

            badge = ''
            if r.get('estado') == 'Vendido':
                badge = 'vendido'
            elif r.get('nuevo_precio'):
                badge = 'nuevo_precio'
            else:
                try:
                    fi_raw = r.get('fecha_ingreso', '')
                    fi_str = fi_raw.strftime('%Y-%m-%d') if hasattr(fi_raw,'strftime') else str(fi_raw)[:10]
                    if fi_str and (ahora - datetime.strptime(fi_str, '%Y-%m-%d')).days <= 7:
                        badge = 'recien_llegado'
                except: pass

            public.append({
                'id':             r.get('id'),
                'marca':          r.get('marca', ''),
                'modelo':         r.get('modelo', ''),
                'anio':           r.get('anio'),
                'tipo_vehiculo':  r.get('tipo_vehiculo', ''),
                'color':          r.get('color', ''),
                'transmision':    r.get('transmision', ''),
                'traccion':       r.get('traccion', ''),
                'combustible':    r.get('combustible', ''),
                'km_aprox':       r.get('km_aprox'),
                'cantidad_duenos':r.get('cantidad_duenos'),
                'motor':          r.get('motor', ''),
                'patente':        r.get('patente', ''),
                'tipo_registro':  r.get('tipo_registro', ''),
                'estado':         r.get('estado', ''),
                'fecha_ingreso':  str(r.get('fecha_ingreso', ''))[:10],
                'precio_vitrina': precio_vitrina,
                'descripcion_publica': r.get('descripcion_publica', ''),
                'foto_portada':   r.get('foto_portada', ''),
                'total_fotos':    int(r.get('total_fotos', 0)),
                'badge':          badge,
            })
        except Exception as e:
            print(f'Error procesando auto {r.get("id","?")}: {e}')
            continue

    return jsonify(public)


import requests as http_requests

GOOGLE_AUTH_URL  = 'https://accounts.google.com/o/oauth2/v2/auth'
GOOGLE_TOKEN_URL = 'https://oauth2.googleapis.com/token'
GOOGLE_CAL_URL   = 'https://www.googleapis.com/calendar/v3/calendars/primary/events'
GOOGLE_SCOPES    = 'https://www.googleapis.com/auth/calendar.events'

@app.route('/oauth/google')
@login_required
def google_oauth_start():
    """Inicia el flujo OAuth con Google Calendar."""
    client_id    = os.environ.get('GOOGLE_CLIENT_ID', '')
    redirect_uri = os.environ.get('GOOGLE_REDIRECT_URI',
                   request.host_url.rstrip('/') + '/oauth/google/callback')
    from urllib.parse import urlencode
    params = {
        'client_id':     client_id,
        'redirect_uri':  redirect_uri,
        'response_type': 'code',
        'scope':         GOOGLE_SCOPES,
        'access_type':   'offline',
        'prompt':        'consent',
        'login_hint':    'contacto@corze.cl',
    }
    url = GOOGLE_AUTH_URL + '?' + urlencode(params)
    return redirect(url)

@app.route('/oauth/google/callback')
@login_required
def google_oauth_callback():
    """Google redirige aquí con el código de autorización."""
    code         = request.args.get('code')
    redirect_uri = os.environ.get('GOOGLE_REDIRECT_URI',
                   request.host_url.rstrip('/') + '/oauth/google/callback')
    if not code:
        return 'Error: no se recibió código de autorización', 400

    resp = http_requests.post(GOOGLE_TOKEN_URL, data={
        'code':          code,
        'client_id':     os.environ.get('GOOGLE_CLIENT_ID'),
        'client_secret': os.environ.get('GOOGLE_CLIENT_SECRET'),
        'redirect_uri':  redirect_uri,
        'grant_type':    'authorization_code',
    })
    tokens = resp.json()
    if 'access_token' not in tokens:
        return f'Error obteniendo token: {tokens}', 400

    session['google_access_token']  = tokens.get('access_token')
    session['google_refresh_token'] = tokens.get('refresh_token', '')
    return redirect(url_for('calendario') + '?calendar=conectado')

@app.route('/api/calendar/disconnect', methods=['POST'])
@login_required
def api_calendar_disconnect():
    session.pop('google_access_token', None)
    session.pop('google_refresh_token', None)
    return jsonify({'ok': True})

def _google_refresh_token():
    """Renueva el access token usando el refresh token guardado en sesión."""
    refresh_token = session.get('google_refresh_token')
    if not refresh_token:
        return False
    try:
        resp = http_requests.post(GOOGLE_TOKEN_URL, data={
            'client_id':     os.environ.get('GOOGLE_CLIENT_ID'),
            'client_secret': os.environ.get('GOOGLE_CLIENT_SECRET'),
            'refresh_token': refresh_token,
            'grant_type':    'refresh_token',
        })
        data = resp.json()
        if 'access_token' in data:
            session['google_access_token'] = data['access_token']
            return True
    except Exception as e:
        print(f'Error renovando token Google: {e}')
    return False

def _google_get(url, params=None):
    """GET a Google API renovando token si es necesario."""
    token = session.get('google_access_token')
    if not token:
        return None, 401
    resp = http_requests.get(url, headers={'Authorization': f'Bearer {token}'}, params=params)
    if resp.status_code == 401 and _google_refresh_token():
        token = session.get('google_access_token')
        resp  = http_requests.get(url, headers={'Authorization': f'Bearer {token}'}, params=params)
    return resp, resp.status_code

def _google_post(url, json_data):
    """POST a Google API renovando token si es necesario."""
    token = session.get('google_access_token')
    if not token:
        return None, 401
    headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
    resp    = http_requests.post(url, headers=headers, json=json_data)
    if resp.status_code == 401 and _google_refresh_token():
        token   = session.get('google_access_token')
        headers['Authorization'] = f'Bearer {token}'
        resp    = http_requests.post(url, headers=headers, json=json_data)
    return resp, resp.status_code

@app.route('/api/calendar/status')
@login_required
def api_calendar_status():
    connected = bool(session.get('google_access_token') or session.get('google_refresh_token'))
    # Si hay refresh token pero no access token, renovar automáticamente
    if session.get('google_refresh_token') and not session.get('google_access_token'):
        connected = _google_refresh_token()
    return jsonify({'connected': connected})

@app.route('/api/calendar/agendar', methods=['POST'])
@login_required
def api_calendar_agendar():
    """Crea un evento en Google Calendar."""
    if not session.get('google_access_token'):
        if not _google_refresh_token():
            return jsonify({'ok': False, 'error': 'Google Calendar no conectado',
                            'auth_url': url_for('google_oauth_start')})

    data       = request.get_json()
    titulo     = data.get('titulo', 'Visita Corze')
    fecha      = data.get('fecha')        # YYYY-MM-DD
    hora_ini   = data.get('hora_ini', '10:00')
    hora_fin   = data.get('hora_fin', '11:00')
    descripcion= data.get('descripcion', '')
    email_inv  = data.get('email_cliente', '')
    ubicacion  = data.get('ubicacion', 'Corze, Santiago, Chile')

    evento = {
        'summary':     titulo,
        'location':    ubicacion,
        'description': descripcion,
        'start': {'dateTime': f'{fecha}T{hora_ini}:00',
                  'timeZone': 'America/Santiago'},
        'end':   {'dateTime': f'{fecha}T{hora_fin}:00',
                  'timeZone': 'America/Santiago'},
        'reminders': {
            'useDefault': False,
            'overrides': [
                {'method': 'email',  'minutes': 60},
                {'method': 'popup',  'minutes': 30},
            ]
        },
    }
    if email_inv:
        evento['attendees'] = [{'email': email_inv}]

    resp, status = _google_post(GOOGLE_CAL_URL, evento)
    if resp is None:
        return jsonify({'ok': False, 'error': 'No conectado',
                        'auth_url': url_for('google_oauth_start')})
    if status in (200, 201):
        ev = resp.json()
        return jsonify({'ok': True, 'event_id': ev.get('id'),
                        'link': ev.get('htmlLink')})
    elif status == 401:
        session.pop('google_access_token', None)
        session.pop('google_refresh_token', None)
        return jsonify({'ok': False, 'error': 'Sesión expirada — reconecta Google Calendar',
                        'auth_url': url_for('google_oauth_start')})
    else:
        return jsonify({'ok': False, 'error': resp.text})

@app.route('/api/calendar/eventos')
@login_required
def api_calendar_eventos():
    """Lista próximos eventos del calendario."""
    token = session.get('google_access_token')
    if not token:
        return jsonify({'ok': False, 'eventos': []})
    ahora = datetime.now(_TZ).isoformat()
    resp, status = _google_get(GOOGLE_CAL_URL, params={
        'timeMin':      ahora,
        'maxResults':   20,
        'orderBy':      'startTime',
        'singleEvents': True,
    })
    if resp and status == 200:
        return jsonify({'ok': True, 'eventos': resp.json().get('items', [])})
    return jsonify({'ok': False, 'eventos': []})

# ══════════════════════════════════════════════════════════════════════════════
#  IA DESCRIPCIÓN — Genera texto de venta con Anthropic API
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/inventario/<doc_id>/generar-descripcion', methods=['POST'])
@login_required
def api_generar_descripcion(doc_id):
    car  = db.get_inventario_by_id(doc_id)
    if not car:
        return jsonify({'ok': False, 'error': 'Vehículo no encontrado'})
    tono = (request.get_json() or {}).get('tono', 'profesional')
    prompt = f"""Genera una descripción de venta atractiva y concisa (máximo 180 palabras) para este vehículo.
Tono: {tono}. Sin encabezados, solo el texto.

- Marca: {car.get('marca','')} | Modelo: {car.get('modelo','')} | Año: {car.get('anio','')}
- Color: {car.get('color','')} | Transmisión: {car.get('transmision','')} | Combustible: {car.get('combustible','')}
- Kilómetros: {car.get('km_aprox','')} | Tracción: {car.get('traccion','')} | Motor: {car.get('motor','')}
- N° dueños anteriores: {car.get('cantidad_duenos','')}
- Notas (contexto, NO mencionar textualmente): {car.get('notas','')}

Resalta puntos fuertes. Menciona que acepta parte de pago si aplica. Termina con llamada a la acción."""
    txt, err = _claude_call(
        prompt,
        system='Eres redactor experto en ventas de vehículos para Corze, Chile.',
        max_tokens=400,
    )
    if err:
        return jsonify({'ok': False, 'error': err})
    return jsonify({'ok': True, 'descripcion': txt.strip()})

# ══════════════════════════════════════════════════════════════════════════════
#  CRM KANBAN — Pipeline de etapas
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/crm/leads', methods=['POST'])
@login_required
def api_crear_lead():
    """Crea un lead manualmente desde el panel admin."""
    d = request.get_json()
    if not d.get('nombre') and not d.get('telefono'):
        return jsonify({'ok': False, 'error': 'Nombre o teléfono requerido'})
    lead = db.create_lead(
        nombre   = d.get('nombre', ''),
        telefono = d.get('telefono', ''),
        email    = d.get('email', ''),
        canal    = d.get('canal', 'manual'),
        vehiculo_interes = d.get('vehiculo_interes', ''),
        notas    = d.get('notas', ''),
        etapa    = d.get('etapa', 'Nuevo'),
        origen   = 'admin',
    )
    return jsonify({'ok': bool(lead.get('id')), 'id': lead.get('id','')})

@app.route('/api/contacto', methods=['POST'])
def api_contacto_vitrina():
    """Lead desde formulario público de la vitrina — sin login requerido."""
    d = request.get_json() or {}
    nombre   = d.get('nombre', '').strip()
    telefono = d.get('telefono', '').strip()
    if not nombre or not telefono:
        return jsonify({'ok': False, 'error': 'Nombre y teléfono son requeridos'})
    lead = db.create_lead(
        nombre   = nombre,
        telefono = telefono,
        email    = d.get('email', ''),
        canal    = 'vitrina',
        vehiculo_interes = d.get('vehiculo', ''),
        notas    = d.get('mensaje', ''),
        etapa    = 'Nuevo',
        origen   = 'vitrina_web',
    )
    return jsonify({'ok': bool(lead.get('id'))})

# ══════════════════════════════════════════════════════════════════════════════
#  TAREAS — Checklist de vehículos y pendientes
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/admin/tareas')
@login_required
def tareas():
    trabajadores = db.get_all_trabajadores(solo_activos=True)
    return render_template('tareas.html', page='tareas', trabajadores=trabajadores)

@app.route('/api/tareas', methods=['GET'])
@login_required
def api_tareas_list():
    estado   = request.args.get('estado', '')
    asignado = request.args.get('asignado', '')
    rows     = db.get_all_tareas(estado=estado, asignado=asignado)
    return jsonify(rows)

@app.route('/api/tareas', methods=['POST'])
@login_required
def api_tareas_add():
    data = request.get_json()
    data['creado_por'] = session.get('usuario', '')
    ok, err = db.add_tarea(data)
    return jsonify({'ok': ok, 'id': err if ok else None, 'error': None if ok else err})

@app.route('/api/tareas/<doc_id>', methods=['PUT'])
@login_required
def api_tareas_update(doc_id):
    data = request.get_json()
    ok, err = db.update_tarea(doc_id, data)
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/tareas/<doc_id>', methods=['DELETE'])
@login_required
def api_tareas_delete(doc_id):
    ok = db.delete_tarea(doc_id)
    return jsonify({'ok': ok})

@app.route('/api/tareas/<doc_id>/toggle', methods=['POST'])
@login_required
def api_tareas_toggle(doc_id):
    data       = request.get_json()
    completada = bool(data.get('completada', False))
    ok = db.toggle_tarea(doc_id, completada, session.get('usuario', ''))
    return jsonify({'ok': ok})

@app.route('/api/tareas/badge')
@login_required
def api_tareas_badge():
    count = db.get_tareas_urgentes_count()
    return jsonify({'urgentes': count})

@app.route('/api/tareas/upcoming')
@login_required
def api_tareas_upcoming():
    """Tareas pendientes dentro de su ventana de recordatorio (por usuario en sesión)."""
    from datetime import timedelta
    today   = _cl().date()
    usuario = session.get('usuario', '')
    docs    = db.get_all_tareas(estado='pendiente')
    upcoming = []
    for t in docs:
        asig = t.get('asignado_a', 'Todos')
        if asig != 'Todos' and asig != usuario:
            continue
        fecha_str = str(t.get('fecha_limite', ''))[:10]
        if not fecha_str:
            continue
        try:
            fecha      = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            recordatorio = int(t.get('recordatorio', 1) or 1)
            delta = (fecha - today).days
            if delta <= 0:
                upcoming.append({**t, 'urgencia': 'vencida', 'delta': delta})
            elif delta <= recordatorio:
                upcoming.append({**t, 'urgencia': f'{delta} día{"s" if delta!=1 else ""}', 'delta': delta})
        except Exception:
            pass
    upcoming.sort(key=lambda x: x.get('delta', 0))
    return jsonify({'tasks': upcoming[:10]})

# ══════════════════════════════════════════════════════════════════════════════
#  CALENDARIO
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/admin/calendario')
@login_required
def calendario():
    trabajadores = db.get_all_trabajadores(solo_activos=True)
    return render_template('calendario.html', page='calendario', trabajadores=trabajadores)

@app.route('/api/calendario', methods=['GET'])
@login_required
def api_calendario_list():
    mes  = request.args.get('mes', '')
    # Eventos propios
    evs  = db.get_eventos_calendario(mes=mes)
    # Tareas como eventos (fecha_limite)
    tareas = db.get_all_tareas(estado='pendiente')
    for t in tareas:
        f = str(t.get('fecha_limite', ''))[:10]
        if not f: continue
        if mes and not f.startswith(mes): continue
        evs.append({
            'id':          t['id'],
            'titulo':      t.get('titulo', ''),
            'fecha':       f,
            'hora_inicio': '',
            'tipo':        'tarea',
            'asignado_a':  t.get('asignado_a', ''),
            'prioridad':   t.get('prioridad', 'media'),
            '_source':     'tarea',
        })
    return jsonify(evs)

@app.route('/api/calendario', methods=['POST'])
@login_required
def api_calendario_add():
    data = request.get_json()
    data['creado_por'] = session.get('usuario', '')
    ok, err = db.add_evento_calendario(data)
    return jsonify({'ok': ok, 'id': err if ok else None, 'error': None if ok else err})

@app.route('/api/calendario/<doc_id>', methods=['PUT'])
@login_required
def api_calendario_update(doc_id):
    data = request.get_json()
    ok, err = db.update_evento_calendario(doc_id, data)
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/calendario/<doc_id>', methods=['DELETE'])
@login_required
def api_calendario_delete(doc_id):
    ok = db.delete_evento_calendario(doc_id)
    return jsonify({'ok': ok})

# ══════════════════════════════════════════════════════════════════════════════
#  BUSCADOR — MercadoLibre + multi-plataforma
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/admin/buscador')
@login_required
def buscador():
    return render_template('buscador.html', page='buscador')

@app.route('/api/buscador/ml', methods=['GET'])
@login_required
def api_buscador_ml():
    """Búsqueda en MercadoLibre Chile — API pública, sin login."""
    q          = request.args.get('q', '').strip()
    precio_min = request.args.get('precio_min', '')
    precio_max = request.args.get('precio_max', '')
    anio_min   = request.args.get('anio_min', '')
    anio_max   = request.args.get('anio_max', '')
    offset     = int(request.args.get('offset', 0))

    if not q and not precio_min and not precio_max:
        return jsonify({'ok': False, 'error': 'Ingresa al menos un criterio', 'results': [], 'total': 0})

    params = {'category': 'MLC1744', 'limit': 24, 'offset': offset}
    if q: params['q'] = q

    price_filter = ''
    if precio_min and precio_max: price_filter = f'{precio_min}-{precio_max}'
    elif precio_min: price_filter = f'{precio_min}-999999999'
    elif precio_max: price_filter = f'0-{precio_max}'
    if price_filter: params['price'] = price_filter

    try:
        resp = http_requests.get('https://api.mercadolibre.com/sites/MLC/search',
                                 params=params, timeout=12)
        data = resp.json()
        results = []
        for r in data.get('results', []):
            attrs = {a['id']: a.get('value_name', '') for a in r.get('attributes', [])}
            anio  = attrs.get('VEHICLE_YEAR', '')
            km    = attrs.get('KILOMETERS', '')

            # Filtros de año locales (ML no siempre los soporta como param)
            if anio_min and anio and int(anio) < int(anio_min): continue
            if anio_max and anio and int(anio) > int(anio_max): continue

            results.append({
                'id':        r.get('id'),
                'titulo':    r.get('title', ''),
                'precio':    int(r.get('price', 0)),
                'moneda':    r.get('currency_id', 'CLP'),
                'thumbnail': (r.get('thumbnail', '') or '').replace('http://', 'https://'),
                'link':      r.get('permalink', ''),
                'año':       anio,
                'km':        km,
                'condicion': attrs.get('ITEM_CONDITION', ''),
                'ubicacion': r.get('address', {}).get('city_name', '') or
                             r.get('seller_address', {}).get('city', {}).get('name', ''),
            })
        return jsonify({'ok': True, 'total': data.get('paging', {}).get('total', 0), 'results': results})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'results': [], 'total': 0})

@app.route('/api/buscador/fb/config', methods=['GET'])
@login_required
def api_fb_config_get():
    cfg = db.get_fb_config()
    return jsonify({'ok': True,
                    'has_cookies': bool(cfg.get('cookies', '').strip()),
                    'updated_at':  cfg.get('updated_at', '')})

@app.route('/api/buscador/fb/config', methods=['POST'])
@login_required
def api_fb_config_save():
    data = request.get_json()
    ok = db.save_fb_config({
        'cookies':    data.get('cookies', '').strip(),
        'updated_por': session.get('usuario', ''),
    })
    return jsonify({'ok': ok})

@app.route('/api/buscador/fb', methods=['GET'])
@login_required
def api_buscador_fb():
    """Scraping de Facebook Marketplace Chile via Playwright."""
    from scraper_fb import scrape, PLAYWRIGHT_OK
    if not PLAYWRIGHT_OK:
        return jsonify({'ok': False,
                        'error': 'Playwright no disponible en este servidor. '
                                 'Asegúrate de que el deploy en Railway incluyó '
                                 '"playwright install chromium --with-deps".',
                        'items': []})
    q          = request.args.get('q', '').strip()
    precio_min = request.args.get('precio_min') or None
    precio_max = request.args.get('precio_max') or None
    anio_min   = request.args.get('anio_min')   or None
    anio_max   = request.args.get('anio_max')   or None
    ciudad     = request.args.get('ciudad', 'santiago').strip() or 'santiago'

    cfg = db.get_fb_config()
    result = scrape(
        query=q, precio_min=precio_min, precio_max=precio_max,
        anio_min=anio_min, anio_max=anio_max,
        cookies_str=cfg.get('cookies', ''),
        ciudad=ciudad,
    )
    # Log debug info en servidor para diagnosticar
    dbg = result.get('debug', {})
    print(f'[FB] url={dbg.get("page_url","")} title={dbg.get("page_title","")} '
          f'graphql={dbg.get("graphql_responses",0)} dom={dbg.get("dom_links",0)} '
          f'items={len(result.get("items",[]))} error={result.get("error")}')
    return jsonify({'ok': not result.get('error'), **result})

@app.route('/api/buscador/guardados', methods=['GET'])
@login_required
def api_buscador_guardados():
    return jsonify(db.get_guardados_buscador())

@app.route('/api/buscador/guardados', methods=['POST'])
@login_required
def api_buscador_guardar():
    data = request.get_json()
    data['guardado_por'] = session.get('usuario', '')
    ok, err = db.add_guardado_buscador(data)
    return jsonify({'ok': ok, 'id': err if ok else None})

@app.route('/api/buscador/guardados/<doc_id>', methods=['DELETE'])
@login_required
def api_buscador_eliminar(doc_id):
    ok = db.delete_guardado_buscador(doc_id)
    return jsonify({'ok': ok})

ETAPAS_CRM = ['Nuevo','Contactado','Agendado','Fotografiado','En Venta','Vendido','Descartado']

@app.route('/api/crm/kanban')
@login_required
def api_crm_kanban():
    """Retorna todas las conversaciones agrupadas por etapa."""
    try:
        convs = db.get_conversaciones()
        grupos = {e: [] for e in ETAPAS_CRM}
        for c in convs:
            etapa = c.get('etapa_crm', 'Nuevo')
            if etapa not in grupos:
                etapa = 'Nuevo'
            grupos[etapa].append({
                'id':             c.get('id'),
                'nombre':         c.get('nombre_cliente') or c.get('telefono','Sin nombre'),
                'telefono':       c.get('telefono',''),
                'canal':          c.get('canal','whatsapp'),
                'vehiculo':       c.get('vehiculo_interes',''),
                'asignado':       c.get('asignado_a',''),
                'ultima_hora':    str(c.get('ultima_hora',''))[:16],
                'etapa_crm':      etapa,
            })
        return jsonify({'ok': True, 'etapas': ETAPAS_CRM, 'grupos': grupos})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/crm/conversacion/<conv_id>/etapa', methods=['PUT'])
@login_required
def api_set_etapa(conv_id):
    data  = request.get_json()
    etapa = data.get('etapa', 'Nuevo')
    if etapa not in ETAPAS_CRM:
        return jsonify({'ok': False, 'error': 'Etapa inválida'})
    ok = db.update_conversacion(conv_id, {'etapa_crm': etapa})
    return jsonify({'ok': ok})


# ══════════════════════════════════════════════════════════════════════════════
#  P&L — Rentabilidad por vehículo
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/inventario/<doc_id>/pnl')
@login_required
def api_inventario_pnl(doc_id):
    car = db.get_inventario_by_id(doc_id)
    if not car:
        return jsonify({'ok': False, 'error': 'No encontrado'})
    transacciones = db.get_finanzas_by_vehiculo(doc_id)
    precio_compra = int(car.get('precio_compra') or car.get('precio_pedido') or 0)
    precio_venta  = int(car.get('precio_venta_final') or 0)
    gastos  = sum(int(t.get('monto', 0) or 0) for t in transacciones if t.get('tipo') == 'costo')
    ingresos_extra = sum(int(t.get('monto', 0) or 0) for t in transacciones if t.get('tipo') == 'ingreso')
    ganancia_neta = (precio_venta + ingresos_extra) - precio_compra - gastos
    return jsonify({
        'ok': True,
        'precio_compra':   precio_compra,
        'precio_venta':    precio_venta,
        'gastos_vinculados': gastos,
        'ingresos_extra':  ingresos_extra,
        'ganancia_neta':   ganancia_neta,
        'transacciones':   transacciones,
        'tipo_registro':   car.get('tipo_registro', ''),
        'comision_monto':  int(car.get('comision_monto') or 0),
        'comision_pct':    float(car.get('comision_porcentaje') or 5),
    })


# ══════════════════════════════════════════════════════════════════════════════
#  HISTORIAL — Log de cambios por vehículo
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/inventario/<doc_id>/historial')
@login_required
def api_inventario_historial(doc_id):
    logs = db.get_historial_vehiculo(doc_id)
    return jsonify(logs)


# ══════════════════════════════════════════════════════════════════════════════
#  ALERTAS DOCUMENTOS — Docs próximos a vencer
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/inventario/alertas/documentos')
@login_required
def api_alertas_documentos():
    dias = int(request.args.get('dias', 30))
    alertas = db.get_docs_por_vencer(dias=dias)
    return jsonify({'alertas': alertas, 'total': len(alertas)})


# ══════════════════════════════════════════════════════════════════════════════
#  COTIZADOR PDF — Cotización imprimible por vehículo
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/admin/inventario/<doc_id>/cotizacion')
@login_required
def cotizacion_print(doc_id):
    car = db.get_inventario_by_id(doc_id)
    if not car:
        return '<h2>Vehículo no encontrado</h2>', 404
    precio = int(request.args.get('precio') or
                 car.get('precio_vitrina') or
                 car.get('precio_venta_colaboradores') or
                 car.get('precio_pedido') or 0)
    validez_dias = int(request.args.get('validez', 7))
    from datetime import timedelta
    validez_fecha = (_cl() + timedelta(days=validez_dias)).strftime('%d/%m/%Y')
    return render_template('cotizacion.html',
        car=car, precio=precio,
        validez_fecha=validez_fecha,
        validez_dias=validez_dias,
        logo_b64=get_logo_b64(),
        fecha_emision=_cl().strftime('%d/%m/%Y %H:%M'),
        agente=session.get('usuario', ''),
        folio=f'COT-{_cl().strftime("%Y%m%d-%H%M%S")}',
    )


# ══════════════════════════════════════════════════════════════════════════════
#  REPORTE MENSUAL PDF — Finanzas
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/admin/finanzas/reporte')
@login_required
def finanzas_reporte_print():
    mes = request.args.get('mes', _cl().strftime('%Y-%m'))
    transacciones = db.get_all_finanzas(mes=mes)
    # Quitar base64 del comprobante para no inflar el HTML
    for t in transacciones:
        t.pop('comprobante', None)
    resumen = db.get_finanzas_resumen(mes=mes)
    return render_template('finanzas_reporte.html',
        mes=mes,
        transacciones=transacciones,
        resumen=resumen,
        logo_b64=get_logo_b64(),
        fecha_generacion=_cl().strftime('%d/%m/%Y %H:%M'),
        agente=session.get('usuario', ''),
    )


# ══════════════════════════════════════════════════════════════════════════════
#  PWA — Service Worker y Manifest
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/sw.js')
def service_worker():
    from flask import send_from_directory
    return send_from_directory('static', 'sw.js',
        mimetype='application/javascript')

@app.route('/manifest.json')
def manifest():
    from flask import send_from_directory
    return send_from_directory('static', 'manifest.json',
        mimetype='application/manifest+json')


# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURACIÓN — Metas / Costos Fijos / WhatsApp Templates
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/admin/config')
@login_required
def config_page():
    return render_template('config.html', page='config')

# ── Metas ──────────────────────────────────────────────────────────────────
@app.route('/api/config/meta', methods=['GET', 'POST'])
@login_required
def api_meta():
    if request.method == 'GET':
        mes = request.args.get('mes', _cl().strftime('%Y-%m'))
        return jsonify(db.get_meta(mes))
    data = request.get_json() or {}
    mes  = data.get('mes', _cl().strftime('%Y-%m'))
    ok   = db.save_meta(mes, int(data.get('unidades', 0)), int(data.get('ingresos', 0)))
    return jsonify({'ok': ok})

# ── Costos fijos ──────────────────────────────────────────────────────────
@app.route('/api/config/costos-fijos', methods=['GET', 'POST'])
@login_required
def api_costos_fijos():
    if request.method == 'GET':
        return jsonify(db.get_costos_fijos())
    data = request.get_json() or {}
    ok   = db.save_costos_fijos(data.get('items', []))
    return jsonify({'ok': ok})

# ── WhatsApp templates ────────────────────────────────────────────────────
@app.route('/api/config/wa-templates', methods=['GET', 'POST'])
@login_required
def api_wa_templates():
    if request.method == 'GET':
        return jsonify(db.get_wa_templates())
    data = request.get_json() or {}
    ok   = db.save_wa_templates(data.get('templates', []))
    return jsonify({'ok': ok})

# ── Precio de mercado (scraper Chileautos) ────────────────────────────────
def _scrape_chileautos(marca, modelo, anio):
    import re
    from urllib.parse import quote
    from playwright.sync_api import sync_playwright

    q   = f"{marca} {modelo} {anio}".strip()
    url = f"https://www.chileautos.cl/vehiculos/?q={quote(q)}"
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=['--no-sandbox', '--disable-dev-shm-usage', '--disable-gpu']
            )
            ctx  = browser.new_context(
                user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
                locale='es-CL',
            )
            page = ctx.new_page()
            page.goto(url, wait_until='domcontentloaded', timeout=20000)
            page.wait_for_timeout(4000)   # tiempo para que React hidrate

            # Extraer precios con JS desde el DOM ya renderizado
            raw_prices = page.evaluate(r"""() => {
                const seen = new Set();
                document.querySelectorAll('*').forEach(el => {
                    const t = el.childNodes.length === 1 && el.firstChild.nodeType === 3
                        ? el.textContent.trim() : '';
                    const m = t.match(/^\$?\s*([\d]{1,3}(?:\.[\d]{3})+)\s*$/);
                    if (m) seen.add(parseInt(m[1].replace(/\./g, '')));
                });
                return Array.from(seen);
            }""")
            browser.close()

        prices = [n for n in raw_prices if 2_000_000 < n < 500_000_000]
        if not prices:
            return None

        prices.sort()
        cut    = max(1, len(prices) // 10)
        prices = prices[cut:-cut] if len(prices) > 5 else prices

        return {
            'count':    len(prices),
            'minimo':   prices[0],
            'maximo':   prices[-1],
            'mediana':  prices[len(prices) // 2],
            'promedio': int(sum(prices) / len(prices)),
            'url':      url,
        }
    except Exception as e:
        print(f'Chileautos scraper error: {e}')
        return None


@app.route('/api/inventario/precio-mercado')
@login_required
def api_precio_mercado():
    from urllib.parse import quote
    marca  = request.args.get('marca', '').strip()
    modelo = request.args.get('modelo', '').strip()
    anio   = request.args.get('anio', '').strip()

    if not marca or not modelo:
        return jsonify({'ok': False, 'error': 'Faltan datos del vehículo'})

    cache_key = f"{marca}_{modelo}_{anio}".lower().replace(' ', '_')
    cached = db.get_precio_mercado_cache(cache_key)
    if cached and cached.get('count'):
        cached['cached'] = True
        return jsonify({'ok': True, **cached})

    result = _scrape_chileautos(marca, modelo, anio)
    if result:
        db.set_precio_mercado_cache(cache_key, result)
        return jsonify({'ok': True, **result})

    # Fallback: links de búsqueda
    q = quote(f"{marca} {modelo} {anio}".strip())
    return jsonify({
        'ok':       True,
        'fallback': True,
        'urls': {
            'chileautos': f"https://www.chileautos.cl/vehiculos/?q={q}",
            'yapo':       f"https://www.yapo.cl/region-metropolitana-de-santiago/autos-usados/?ca=33_s&keywords={q}",
        },
    })


# ══════════════════════════════════════════════════════════════════════════════
#  CLAUDE AI — helper + rutas
# ══════════════════════════════════════════════════════════════════════════════

def _claude_call(prompt, system=None, max_tokens=1024):
    """Llama a Claude Haiku. Retorna (texto, error)."""
    api_key = db.get_anthropic_key() or os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return None, 'API key de Claude no configurada — agrégala en ⚙️ Configuración > IA'
    body = {
        'model':      'claude-haiku-4-5-20251001',
        'max_tokens': max_tokens,
        'messages':   [{'role': 'user', 'content': prompt}],
    }
    if system:
        body['system'] = system
    try:
        r = http_requests.post(
            'https://api.anthropic.com/v1/messages',
            headers={
                'x-api-key':         api_key,
                'anthropic-version': '2023-06-01',
                'content-type':      'application/json',
            },
            json=body,
            timeout=40,
        )
        data = r.json()
        if 'content' in data:
            return data['content'][0]['text'], None
        return None, data.get('error', {}).get('message', str(data))
    except Exception as e:
        return None, str(e)


# ── Config Claude/Anthropic key ────────────────────────────────────────────
@app.route('/api/config/claude', methods=['GET', 'POST'])
@login_required
def api_claude_config():
    if request.method == 'GET':
        key = db.get_anthropic_key() or os.environ.get('ANTHROPIC_API_KEY', '')
        masked = ('*' * (len(key) - 8) + key[-8:]) if len(key) > 8 else ('*' * len(key) if key else '')
        return jsonify({'has_key': bool(key), 'masked': masked})
    data = request.get_json() or {}
    ok = db.save_anthropic_key(data.get('api_key', ''))
    return jsonify({'ok': ok})


# ── Test Claude connection ─────────────────────────────────────────────────
@app.route('/api/config/claude/test', methods=['POST'])
@login_required
def api_claude_test():
    txt, err = _claude_call('Responde solo con: "OK — Claude conectado correctamente"')
    if err:
        return jsonify({'ok': False, 'error': err})
    return jsonify({'ok': True, 'respuesta': txt.strip()})


# ── Descripción generada por Claude ───────────────────────────────────────
@app.route('/api/ai/descripcion', methods=['POST'])
@login_required
def api_ai_descripcion():
    data   = request.get_json() or {}
    doc_id = data.get('doc_id', '')
    tono   = data.get('tono', 'profesional')
    car    = db.get_inventario_by_id(doc_id) if doc_id else data
    prompt = f"""Genera una descripción de venta atractiva y concisa (máximo 200 palabras) para este vehículo.
Tono: {tono}. Sin encabezados. Solo el texto corrido.

VEHÍCULO:
- {car.get('marca','')} {car.get('modelo','')} {car.get('anio','')}
- Color: {car.get('color','')} | Transmisión: {car.get('transmision','')} | Combustible: {car.get('combustible','')}
- Km: {car.get('km_aprox','')} | Motor: {car.get('motor','')} | Tracción: {car.get('traccion','')}
- Dueños previos: {car.get('cantidad_duenos','—')}
- Equipamiento/notas (contexto, no mencionar textualmente): {car.get('notas','')}

Resalta puntos fuertes. Termina con llamada a la acción."""
    txt, err = _claude_call(
        prompt,
        system='Eres redactor experto en ventas de vehículos para Corze, Chile.',
        max_tokens=400,
    )
    if err:
        return jsonify({'ok': False, 'error': err})
    return jsonify({'ok': True, 'descripcion': txt.strip()})


# ── Precio recomendado por Claude ─────────────────────────────────────────
@app.route('/api/ai/precio-recomendado', methods=['POST'])
@login_required
def api_ai_precio():
    data   = request.get_json() or {}
    doc_id = data.get('doc_id', '')
    car    = db.get_inventario_by_id(doc_id) if doc_id else data
    prompt = f"""Dado el siguiente vehículo, recomienda un rango de precio de publicación en pesos chilenos (CLP).
Considera el mercado chileno actual, kilometraje, año, condición y número de dueños.

VEHÍCULO:
- Marca: {car.get('marca','')} | Modelo: {car.get('modelo','')} | Año: {car.get('anio','')}
- Kilómetros: {car.get('km_aprox','')} | Transmisión: {car.get('transmision','')}
- Combustible: {car.get('combustible','')} | Tracción: {car.get('traccion','')}
- Color: {car.get('color','')} | Dueños previos: {car.get('cantidad_duenos','—')}
- Estado general: {car.get('notas','')}

Responde SOLO con JSON válido (sin markdown):
{{"precio_minimo":0,"precio_maximo":0,"precio_recomendado":0,"justificacion":"...","factores_positivos":["..."],"factores_negativos":["..."]}}"""

    txt, err = _claude_call(
        prompt,
        system='Eres experto en tasación de vehículos usados en Chile (mercado 2024-2025).',
        max_tokens=600,
    )
    if err:
        return jsonify({'ok': False, 'error': err})
    try:
        import re as _re, json as _json
        m = _re.search(r'\{.*\}', txt, _re.DOTALL)
        if m:
            return jsonify({'ok': True, **_json.loads(m.group())})
    except Exception:
        pass
    return jsonify({'ok': True, 'raw': txt})


# ── Notas de voz ──────────────────────────────────────────────────────────
@app.route('/api/inventario/<doc_id>/notas-voz', methods=['GET'])
@login_required
def api_get_notas_voz(doc_id):
    notas = db.get_notas_voz(doc_id)
    for n in notas:
        n.pop('audio_b64', None)  # no enviar el audio crudo al listar
    return jsonify(notas)


@app.route('/api/inventario/<doc_id>/notas-voz', methods=['POST'])
@login_required
def api_add_nota_voz(doc_id):
    data          = request.get_json() or {}
    transcripcion = data.get('transcripcion', '').strip()
    duracion      = int(data.get('duracion', 0))

    if not transcripcion:
        return jsonify({'ok': False, 'error': 'Sin transcripción'})

    prompt = f"""Esta es la transcripción de una nota de voz de un gestor de automotora sobre un vehículo:

"{transcripcion}"

Por favor:
1. Crea un resumen interpretado con los puntos clave (estado, observaciones, acuerdos, condiciones del propietario, etc.).
2. Extrae las tareas o acciones concretas mencionadas (si las hay).

Responde SOLO con JSON válido (sin markdown):
{{"resumen":"...","tareas":["..."]}}"""

    txt, err = _claude_call(
        prompt,
        system='Eres asistente de gestión para una automotora en Chile. Eres conciso y preciso.',
        max_tokens=500,
    )

    resumen = ''
    tareas  = []
    if not err and txt:
        try:
            import re as _re, json as _json
            m = _re.search(r'\{.*\}', txt, _re.DOTALL)
            if m:
                parsed  = _json.loads(m.group())
                resumen = parsed.get('resumen', '')
                tareas  = parsed.get('tareas', [])
        except Exception:
            resumen = txt

    nota_id = db.add_nota_voz(doc_id, {
        'transcripcion': transcripcion,
        'resumen':       resumen,
        'tareas':        tareas,
        'duracion_seg':  duracion,
        'usuario':       session.get('usuario', ''),
    })
    return jsonify({'ok': True, 'id': nota_id,
                    'transcripcion': transcripcion, 'resumen': resumen, 'tareas': tareas})


@app.route('/api/inventario/<doc_id>/notas-voz/<nota_id>', methods=['DELETE'])
@login_required
def api_delete_nota_voz(doc_id, nota_id):
    ok = db.delete_nota_voz(nota_id)
    return jsonify({'ok': ok})


# ══════════════════════════════════════════════════════════════════════════════
#  BÚSQUEDA DE PATENTE — Consulta datos del vehículo por patente (via 2captcha)
# ══════════════════════════════════════════════════════════════════════════════

import threading as _threading
import uuid as _uuid
import re as _re
import time as _time

_PATENTE_SITE_KEY  = '6Ld-qMkkAAAAAJ4qghlYL20l7-I1z5A3fwW5a3-U'
_PATENTE_SITE_URL  = 'https://patenteschile.cl/'
_PATENTE_SEARCH_URL= 'https://patenteschile.cl/wp-search/searchPatente.php'
_patente_jobs      = {}   # job_id -> {'status': pending|done|error, ...}

def _patente_worker(job_id, patente, api_key):
    """Corre en un thread separado: resuelve captcha y consulta patenteschile.cl."""
    try:
        # 1. Enviar tarea a 2captcha
        r = http_requests.post('https://2captcha.com/in.php', data={
            'key': api_key, 'method': 'userrecaptcha',
            'googlekey': _PATENTE_SITE_KEY, 'pageurl': _PATENTE_SITE_URL, 'json': 1,
        }, timeout=20)
        d = r.json()
        if d.get('status') != 1:
            _patente_jobs[job_id] = {'status': 'error', 'error': f'2captcha rechazó la tarea: {d}'}
            return
        task_id = d['request']
        print(f'[patente:{job_id}] 2captcha task={task_id}')

        # 2. Polling resultado
        token = None
        for i in range(30):
            _time.sleep(5)
            poll = http_requests.get('https://2captcha.com/res.php', params={
                'key': api_key, 'action': 'get', 'id': task_id, 'json': 1,
            }, timeout=15)
            res = poll.json()
            print(f'[patente:{job_id}] poll {(i+1)*5}s → {str(res.get("request",""))[:30]}')
            if res.get('status') == 1:
                token = res['request']
                break
            if res.get('request') not in ('CAPCHA_NOT_READY', 'CAPTCHA_NOT_READY'):
                _patente_jobs[job_id] = {'status': 'error', 'error': f'Error 2captcha: {res}'}
                return

        if not token:
            _patente_jobs[job_id] = {'status': 'error', 'error': 'Tiempo agotado resolviendo captcha'}
            return

        # 3. Consultar patenteschile.cl
        resp = http_requests.post(_PATENTE_SEARCH_URL, data={
            'patente': patente, 'g-recaptcha-response': token,
        }, headers={
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
            'Referer': _PATENTE_SITE_URL, 'Origin': 'https://patenteschile.cl',
        }, timeout=30)
        print(f'[patente:{job_id}] patenteschile status={resp.status_code} len={len(resp.text)} body={resp.text[:150]}')

        if resp.status_code == 500:
            _patente_jobs[job_id] = {'status': 'error', 'error': f'Patente {patente} no encontrada en el registro'}
            return
        if 'alert-danger' in resp.text or ('reCAPTCHA' in resp.text and len(resp.text) < 200):
            _patente_jobs[job_id] = {'status': 'error', 'error': 'Captcha rechazado. Intenta de nuevo.'}
            return

        # 4. Parsear HTML
        datos = _parsear_datos_patente(resp.text)
        if not datos:
            print(f'[patente:{job_id}] Sin datos. HTML:\n{resp.text[:600]}')
            _patente_jobs[job_id] = {'status': 'error', 'error': 'Patente sin datos disponibles en el registro'}
            return

        print(f'[patente:{job_id}] Datos: {datos}')
        _patente_jobs[job_id] = {'status': 'done', 'datos': datos}

    except Exception as exc:
        import traceback as _tb
        print(f'[patente:{job_id}] Excepción: {_tb.format_exc()}')
        _patente_jobs[job_id] = {'status': 'error', 'error': str(exc)}


def _parsear_datos_patente(html):
    clean = lambda s: _re.sub(r'<[^>]+>', '', s).strip()

    def _map(label, value, data):
        l, v = label.lower(), value.strip()
        if not v: return
        if 'marca'      in l: data['marca']      = v
        elif 'modelo'   in l: data['modelo']     = v
        elif 'año'      in l or 'ano' in l or 'fabricaci' in l: data['anio'] = v
        elif 'color'    in l: data['color']      = v
        elif 'tipo'     in l or 'carrocer' in l: data['tipo']  = v
        elif 'motor'    in l: data['motor']      = v
        elif 'chasis'   in l: data['chasis']     = v
        elif 'combustib'in l: data['combustible']= v
        elif 'transmis' in l: data['transmision']= v

    data = {}
    # Formato DataTable: <thead> columnas + <tbody> fila de datos
    thead = _re.search(r'<thead[^>]*>(.*?)</thead>', html, _re.S | _re.I)
    tbody = _re.search(r'<tbody[^>]*>(.*?)</tbody>', html, _re.S | _re.I)
    if thead and tbody:
        headers = [clean(h) for h in _re.findall(r'<th[^>]*>(.*?)</th>', thead.group(1), _re.S | _re.I)]
        first_row = _re.search(r'<tr[^>]*>(.*?)</tr>', tbody.group(1), _re.S | _re.I)
        if first_row:
            cells = [clean(c) for c in _re.findall(r'<td[^>]*>(.*?)</td>', first_row.group(1), _re.S | _re.I)]
            for h, v in zip(headers, cells):
                _map(h, v, data)
    # Fallback: tabla key-value
    if not data:
        for row in _re.findall(r'<tr[^>]*>(.*?)</tr>', html, _re.S | _re.I):
            cells = _re.findall(r'<td[^>]*>(.*?)</td>', row, _re.S | _re.I)
            if len(cells) >= 2:
                _map(clean(cells[0]), clean(cells[1]), data)
    return data


@app.route('/api/buscar-patente', methods=['POST'])
@login_required
def api_buscar_patente():
    """Inicia la búsqueda en background y retorna un job_id inmediatamente."""
    two_captcha_key = os.environ.get('TWO_CAPTCHA_KEY', '').strip()
    if not two_captcha_key:
        return jsonify({'ok': False, 'error': 'TWO_CAPTCHA_KEY no configurada en Railway Variables'})

    raw = (request.get_json(force=True, silent=True) or {}).get('patente', '')
    patente = _re.sub(r'[^A-Z0-9]', '', raw.strip().upper())
    if not patente:
        return jsonify({'ok': False, 'error': 'Patente requerida'})

    job_id = str(_uuid.uuid4())[:8]
    _patente_jobs[job_id] = {'status': 'pending'}
    _threading.Thread(target=_patente_worker, args=(job_id, patente, two_captcha_key), daemon=True).start()
    print(f'[buscar-patente] Job {job_id} iniciado para patente {patente}')
    return jsonify({'ok': True, 'job_id': job_id})


@app.route('/api/buscar-patente/<job_id>', methods=['GET'])
@login_required
def api_buscar_patente_poll(job_id):
    """Retorna el estado actual del job de búsqueda."""
    job = _patente_jobs.get(job_id)
    if not job:
        return jsonify({'ok': False, 'error': 'Job no encontrado'})
    if job['status'] == 'pending':
        return jsonify({'ok': True, 'status': 'pending'})
    if job['status'] == 'done':
        _patente_jobs.pop(job_id, None)  # limpiar
        return jsonify({'ok': True, 'status': 'done', 'datos': job['datos']})
    # error
    _patente_jobs.pop(job_id, None)
    return jsonify({'ok': False, 'status': 'error', 'error': job.get('error', 'Error desconocido')})


# ══════════════════════════════════════════════════════════════════════════════
#  FLIPPING — Gestión de negocios de flipping de vehículos
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/admin/flipping')
@login_required
def flipping():
    from config import MARCAS, TIPOS_VEHICULO, TRANSMISIONES, COMBUSTIBLES
    return render_template('flipping.html', page='flipping',
        marcas=MARCAS, tipos=TIPOS_VEHICULO,
        transmisiones=TRANSMISIONES, combustibles=COMBUSTIBLES,
        users=USERS)

@app.route('/api/flipping', methods=['GET'])
@login_required
def api_flipping_list():
    q            = request.args.get('q', '')
    calificacion = request.args.get('calificacion', '')
    estado       = request.args.get('estado', '')
    rows = db.get_all_flipping(query=q, calificacion=calificacion, estado=estado)
    return jsonify(rows)

@app.route('/api/flipping', methods=['POST'])
@login_required
def api_flipping_add():
    data = request.get_json()
    data['creado_por'] = session.get('usuario', '')
    ok, err = db.add_flipping(data)
    return jsonify({'ok': ok, 'id': err if ok else None, 'error': None if ok else err})

@app.route('/api/flipping/<doc_id>', methods=['GET'])
@login_required
def api_flipping_get(doc_id):
    return jsonify(db.get_flipping_by_id(doc_id) or {})

@app.route('/api/flipping/<doc_id>', methods=['PUT'])
@login_required
def api_flipping_update(doc_id):
    data = request.get_json()
    data['editado_por'] = session.get('usuario', '')
    ok, err = db.update_flipping(doc_id, data)
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/flipping/<doc_id>', methods=['DELETE'])
@login_required
def api_flipping_delete(doc_id):
    ok = db.delete_flipping(doc_id, session.get('usuario', ''))
    return jsonify({'ok': ok})

@app.route('/api/flipping/<doc_id>/simular', methods=['POST'])
@login_required
def api_flipping_simular(doc_id):
    """Guarda los participantes y % de participación del negocio."""
    data         = request.get_json()
    participantes = data.get('participantes', [])
    ok, err = db.update_flipping(doc_id, {
        'participantes': participantes,
        'editado_por': session.get('usuario', ''),
    })
    return jsonify({'ok': ok, 'error': err})


# ══════════════════════════════════════════════════════════════════════════════
#  CAR HUNTER — Gestión de búsquedas de vehículos por cliente
# ══════════════════════════════════════════════════════════════════════════════

# Lista fija de clientes frecuentes (KR Automotora es el principal)
_KR_WISH_LIST = [
    {'marca': 'PEUGEOT',   'modelos': ['208', '2008', '301'],
     'anio_min': None, 'notas': 'Lista KR Automotora'},
    {'marca': 'KIA',       'modelos': ['Río', 'Morning 4', 'Morning 5', 'Soluto', 'Sportage'],
     'anio_min': None, 'notas': 'Lista KR Automotora'},
    {'marca': 'HYUNDAI',   'modelos': ['Accent', 'Grand i10', 'i20', 'Atos'],
     'anio_min': None, 'notas': 'Lista KR Automotora'},
    {'marca': 'SUZUKI',    'modelos': ['Swift', 'Baleno', 'Vitara', 'Nómade'],
     'anio_min': None, 'notas': 'Lista KR Automotora'},
    {'marca': 'CHERY',     'modelos': ['Tiggo 2', 'Tiggo 3', 'Tiggo 7', 'Tiggo 8'],
     'anio_min': None, 'notas': 'Lista KR Automotora'},
    {'marca': 'CHEVROLET', 'modelos': ['Sail', 'Spark', 'Groove'],
     'anio_min': None, 'notas': 'Lista KR Automotora'},
    {'marca': 'CITROEN',   'modelos': ['C-3'],
     'anio_min': 2021,  'notas': 'Lista KR Automotora — 2021 en adelante'},
    {'marca': 'MAZDA',     'modelos': ['Mazda 2', 'Mazda 3'],
     'anio_min': None, 'notas': 'Lista KR Automotora'},
    {'marca': 'MG',        'modelos': ['3', 'ZS', 'ZX', 'GT'],
     'anio_min': None, 'notas': 'Lista KR Automotora'},
]

@app.route('/admin/car-hunter')
@login_required
def car_hunter():
    from config import MARCAS, TIPOS_VEHICULO, TRANSMISIONES, COMBUSTIBLES
    return render_template('car_hunter.html', page='car_hunter',
        marcas=MARCAS, tipos=TIPOS_VEHICULO,
        transmisiones=TRANSMISIONES, combustibles=COMBUSTIBLES,
        users=USERS)

@app.route('/api/carhunter', methods=['GET'])
@login_required
def api_carhunter_list():
    cliente  = request.args.get('cliente', '')
    estado   = request.args.get('estado', '')
    prioridad= request.args.get('prioridad', '')
    q        = request.args.get('q', '')
    rows = db.get_all_carhunter(cliente=cliente, estado=estado, prioridad=prioridad, query=q)
    return jsonify(rows)

@app.route('/api/carhunter', methods=['POST'])
@login_required
def api_carhunter_add():
    data = request.get_json()
    data['creado_por'] = session.get('usuario', '')
    ok, err = db.add_carhunter(data)
    return jsonify({'ok': ok, 'id': err if ok else None, 'error': None if ok else err})

@app.route('/api/carhunter/<doc_id>', methods=['GET'])
@login_required
def api_carhunter_get(doc_id):
    return jsonify(db.get_carhunter_by_id(doc_id) or {})

@app.route('/api/carhunter/<doc_id>', methods=['PUT'])
@login_required
def api_carhunter_update(doc_id):
    data = request.get_json()
    data['editado_por'] = session.get('usuario', '')
    ok, err = db.update_carhunter(doc_id, data)
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/carhunter/<doc_id>', methods=['DELETE'])
@login_required
def api_carhunter_delete(doc_id):
    ok = db.delete_carhunter(doc_id, session.get('usuario', ''))
    return jsonify({'ok': ok})

@app.route('/api/carhunter/seed-kr', methods=['POST'])
@login_required
def api_carhunter_seed_kr():
    """Crea/actualiza la lista KR Automotora sin duplicar marcas ni modelos."""
    # Índice de registros KR existentes por marca (marca → {id, modelos})
    existentes = {
        r['marca'].upper(): r
        for r in db.get_all_carhunter(cliente='KR Automotora')
    }

    creados = 0
    actualizados = 0
    sin_cambios = 0
    errores = []

    for item in _KR_WISH_LIST:
        marca_key = item['marca'].upper()
        modelos_nuevos = item['modelos']

        if marca_key in existentes:
            # Ya existe — agregar solo modelos que no están registrados
            existing = existentes[marca_key]
            modelos_actuales = existing.get('modelos') or []
            a_agregar = [m for m in modelos_nuevos if m not in modelos_actuales]
            if not a_agregar:
                sin_cambios += 1
                continue
            merged = modelos_actuales + a_agregar
            ok, err = db.update_carhunter(existing['id'], {'modelos': merged})
            if ok:
                actualizados += 1
            else:
                errores.append(str(err))
        else:
            # No existe — crear registro completo
            data = {
                'cliente':           'KR Automotora',
                'marca':             item['marca'],
                'modelos':           modelos_nuevos,
                'anio_min':          item.get('anio_min') or 0,
                'anio_max':          0,
                'km_max':            0,
                'duenos_max':        0,
                'colores_preferidos': [],
                'transmision':       '',
                'combustible':       '',
                'presupuesto_max':   0,
                'prioridad':         'Alta',
                'estado':            'Buscando',
                'notas':             item.get('notas', ''),
                'creado_por':        session.get('usuario', ''),
            }
            ok, err = db.add_carhunter(data)
            if ok:
                creados += 1
            else:
                errores.append(str(err))

    return jsonify({
        'ok': True,
        'creados': creados,
        'actualizados': actualizados,
        'sin_cambios': sin_cambios,
        'errores': errores,
    })

@app.route('/api/carhunter/clear-kr', methods=['DELETE'])
@login_required
def api_carhunter_clear_kr():
    """Elimina todos los registros de KR Automotora."""
    registros = db.get_all_carhunter(cliente='KR Automotora')
    eliminados = 0
    for r in registros:
        if db.delete_carhunter(r['id'], session.get('usuario', '')):
            eliminados += 1
    return jsonify({'ok': True, 'eliminados': eliminados})


# ══════════════════════════════════════════════════════════════════════════════
#  INVENTARIO SOLAR
# ══════════════════════════════════════════════════════════════════════════════

CATEGORIAS_SOLAR = [
    'Panel Solar', 'Inversor', 'Batería', 'Estructura/Mounting',
    'Cable/Conector', 'Medidor/Monitoreo', 'Accesorio', 'Kit Completo',
]

@app.route('/admin/inventario-solar')
@login_required
def inventario_solar():
    return render_template('inventario_solar.html', page='inventario_solar',
                           categorias=CATEGORIAS_SOLAR)

@app.route('/api/productos', methods=['GET'])
@login_required
def api_productos_list():
    categoria   = request.args.get('categoria', '')
    q           = request.args.get('q', '')
    solo_activos = request.args.get('solo_activos', 'true').lower() != 'false'
    rows = db.get_all_productos(categoria=categoria, query=q, solo_activos=solo_activos)
    return jsonify(rows)

@app.route('/api/productos/add', methods=['POST'])
@login_required
def api_productos_add():
    data = request.get_json()
    data['creado_por'] = session.get('usuario', '')
    ok, err = db.add_producto(data)
    return jsonify({'ok': ok, 'id': err if ok else None, 'error': None if ok else err})

@app.route('/api/productos/<doc_id>', methods=['GET'])
@login_required
def api_productos_get(doc_id):
    return jsonify(db.get_producto_by_id(doc_id) or {})

@app.route('/api/productos/<doc_id>/edit', methods=['POST'])
@login_required
def api_productos_edit(doc_id):
    data = request.get_json()
    data['editado_por'] = session.get('usuario', '')
    ok, err = db.update_producto(doc_id, data)
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/productos/<doc_id>/delete', methods=['POST'])
@login_required
def api_productos_delete(doc_id):
    ok = db.delete_producto(doc_id)
    return jsonify({'ok': ok})

@app.route('/api/productos/import-excel', methods=['POST'])
@login_required
def api_productos_import_excel():
    try:
        import openpyxl
    except ImportError:
        return jsonify({'ok': False, 'error': 'openpyxl no instalado'}), 500

    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        return jsonify({'ok': False, 'error': 'Archivo .xlsx requerido'})

    try:
        import io
        wb = openpyxl.load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        headers = [str(c.value or '').strip().lower() for c in ws[1]]

        col_map = {
            'codigo':       ['codigo', 'código'],
            'nombre':       ['nombre'],
            'categoria':    ['categoria', 'categoría'],
            'descripcion':  ['descripcion', 'descripción'],
            'marca':        ['marca'],
            'modelo':       ['modelo'],
            'precio_costo': ['precio costo', 'costo', 'precio_costo'],
            'precio_venta': ['precio venta', 'precio_venta', 'venta'],
            'stock_actual': ['stock', 'stock actual', 'stock_actual'],
            'unidad':       ['unidad'],
            'potencia_w':   ['potencia w', 'potencia_w', 'potencia'],
        }

        def get_col(h_list):
            for h in h_list:
                for i, header in enumerate(headers):
                    if h in header:
                        return i
            return None

        idx = {k: get_col(v) for k, v in col_map.items()}

        creados = 0
        actualizados = 0
        errores = []
        usuario = session.get('usuario', '')

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            try:
                def val(k):
                    i = idx.get(k)
                    return str(row[i] or '').strip() if i is not None and i < len(row) else ''

                codigo = val('codigo')
                nombre = val('nombre')
                if not nombre:
                    continue

                data = {
                    'codigo':       codigo,
                    'nombre':       nombre,
                    'categoria':    val('categoria') or 'Accesorio',
                    'descripcion':  val('descripcion'),
                    'marca':        val('marca'),
                    'modelo':       val('modelo'),
                    'precio_costo': int(float(val('precio_costo') or 0)),
                    'precio_venta': int(float(val('precio_venta') or 0)),
                    'stock_actual': int(float(val('stock_actual') or 0)),
                    'stock_minimo': 0,
                    'unidad':       val('unidad') or 'unidad',
                    'potencia_w':   int(float(val('potencia_w') or 0)),
                    'activo':       True,
                    'creado_por':   usuario,
                }

                # Upsert por código si existe
                if codigo:
                    existentes = db.get_all_productos(solo_activos=False)
                    match = next((r for r in existentes if r.get('codigo') == codigo), None)
                    if match:
                        ok, err = db.update_producto(match['id'], data)
                        if ok:
                            actualizados += 1
                        else:
                            errores.append(f'Error actualizando {codigo}: {err}')
                        continue

                ok, err = db.add_producto(data)
                if ok:
                    creados += 1
                else:
                    errores.append(f'Error creando {nombre}: {err}')
            except Exception as e:
                errores.append(f'Fila error: {str(e)}')

        return jsonify({'ok': True, 'creados': creados,
                        'actualizados': actualizados, 'errores': errores})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# ══════════════════════════════════════════════════════════════════════════════
#  PRESUPUESTOS
# ══════════════════════════════════════════════════════════════════════════════

EMAIL_HOST      = os.environ.get('EMAIL_HOST', 'mail.corze.cl')
EMAIL_PORT      = int(os.environ.get('EMAIL_PORT', 587))
EMAIL_USER      = os.environ.get('EMAIL_USER', 'contacto@corze.cl')
EMAIL_PASS      = os.environ.get('EMAIL_PASS', '')
EMAIL_IMAP_HOST = os.environ.get('EMAIL_IMAP_HOST', 'mail.corze.cl')
EMAIL_IMAP_PORT = int(os.environ.get('EMAIL_IMAP_PORT', '993'))

# ── IMAP → Leads ─────────────────────────────────────────────────────────────

def _imap_decode(value):
    if not value:
        return ''
    parts = _email_decode_header(value)
    result = []
    for b, enc in parts:
        if isinstance(b, bytes):
            result.append(b.decode(enc or 'utf-8', errors='ignore'))
        else:
            result.append(b)
    return ' '.join(result)

def _imap_get_body(msg):
    if msg.is_multipart():
        for part in msg.walk():
            if (part.get_content_type() == 'text/plain' and
                    'attachment' not in str(part.get('Content-Disposition', ''))):
                payload = part.get_payload(decode=True)
                if payload:
                    return payload.decode(part.get_content_charset() or 'utf-8', errors='ignore')
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            return payload.decode(msg.get_content_charset() or 'utf-8', errors='ignore')
    return ''

def procesar_email_como_lead(asunto, remitente, cuerpo, usuario='Sistema'):
    """Convierte un email en lead o agrega al historial si ya existe. Retorna (ok, msg, lead_id)."""
    nombre_raw, email_addr = _parseaddr(remitente)
    email_addr = (email_addr or remitente).strip().lower()
    nombre_raw = nombre_raw.strip() or email_addr.split('@')[0]
    partes   = nombre_raw.split()
    nombre   = partes[0] if partes else nombre_raw
    apellido = ' '.join(partes[1:]) if len(partes) > 1 else ''

    existing = db.get_lead_by_email(email_addr) if email_addr else None
    if existing:
        db.add_historial_lead(existing['id'], 'email_recibido', usuario,
                              f'Asunto: {asunto[:120]}')
        return True, f'Email agregado al historial de {nombre}', existing['id']

    notas = f'Asunto: {asunto}\n\n{cuerpo[:800]}' if cuerpo else f'Asunto: {asunto}'
    lead_data = {
        'nombre': nombre, 'apellido': apellido,
        'email': email_addr, 'telefono': '', 'empresa': '',
        'origen': 'Email', 'etapa': 'Nuevo Lead',
        'tipo_proyecto': '', 'asignado_a': '',
        'consumo_kwh': '', 'region': '', 'notas': notas,
    }
    ok, lead_id = db.add_lead(lead_data)
    if ok:
        db.add_historial_lead(lead_id, 'email_recibido', usuario,
                              f'Email detectado: {asunto[:120]}')
    return ok, ('Lead creado desde email' if ok else 'Error creando lead'), (lead_id if ok else None)

def _imap_check_once():
    """Conecta por IMAP SSL y procesa emails no leídos. Retorna cantidad procesada."""
    if not EMAIL_PASS:
        return 0
    processed = 0
    try:
        mail = imaplib.IMAP4_SSL(EMAIL_IMAP_HOST, EMAIL_IMAP_PORT)
        mail.login(EMAIL_USER, EMAIL_PASS)
        mail.select('INBOX')
        _, msgnums = mail.search(None, 'UNSEEN')
        ids = msgnums[0].split() if msgnums[0] else []
        for num in ids:
            try:
                _, data = mail.fetch(num, '(RFC822)')
                raw = data[0][1]
                msg      = _email_lib.message_from_bytes(raw)
                asunto   = _imap_decode(msg.get('Subject', 'Sin asunto'))
                remit    = _imap_decode(msg.get('From', ''))
                cuerpo   = _imap_get_body(msg)
                ok, _, _ = procesar_email_como_lead(asunto, remit, cuerpo)
                if ok:
                    mail.store(num, '+FLAGS', '\\Seen')
                    processed += 1
            except Exception as e:
                print(f'[IMAP] Error procesando email: {e}')
        mail.close()
        mail.logout()
    except Exception as e:
        print(f'[IMAP] Error de conexión: {e}')
    return processed

def _imap_worker():
    import time
    print('[IMAP] Poller iniciado — revisando cada 5 min')
    while True:
        try:
            n = _imap_check_once()
            if n:
                print(f'[IMAP] {n} emails procesados como leads')
        except Exception as e:
            print(f'[IMAP] Worker error: {e}')
        time.sleep(300)

# Iniciar poller en background (seguro con gunicorn --workers 1)
_imap_thread = threading.Thread(target=_imap_worker, daemon=True)
_imap_thread.start()

# API: revisar bandeja ahora (manual trigger)
@app.route('/api/leads/check-emails', methods=['POST'])
@login_required
def api_check_emails():
    try:
        n = _imap_check_once()
        return jsonify({'ok': True, 'procesados': n,
                        'msg': f'{n} email(s) convertidos en leads' if n else 'No hay emails nuevos'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

# API: importar email manualmente
@app.route('/api/leads/importar-email', methods=['POST'])
@login_required
def api_importar_email():
    data     = request.get_json() or {}
    remit    = data.get('remitente', '').strip()
    asunto   = data.get('asunto', '').strip()
    cuerpo   = data.get('cuerpo', '').strip()
    usuario  = session.get('usuario', 'Manual')
    if not remit:
        return jsonify({'ok': False, 'error': 'El campo "De" es requerido'})
    ok, msg, lead_id = procesar_email_como_lead(asunto, remit, cuerpo, usuario)
    return jsonify({'ok': ok, 'msg': msg, 'lead_id': lead_id})

@app.route('/admin/presupuestos')
@login_required
def presupuestos():
    rows = db.get_all_presupuestos()
    return render_template('presupuestos.html', page='presupuestos', presupuestos=rows)

@app.route('/admin/presupuestos/nuevo')
@login_required
def presupuesto_nuevo():
    trabajadores = db.get_all_trabajadores(solo_activos=True)
    return render_template('presupuesto_builder.html', page='presupuestos',
                           presupuesto=None, trabajadores=trabajadores)

@app.route('/admin/presupuestos/<doc_id>')
@login_required
def presupuesto_ver(doc_id):
    p = db.get_presupuesto_by_id(doc_id)
    if not p:
        return redirect(url_for('presupuestos'))
    trabajadores = db.get_all_trabajadores(solo_activos=True)
    return render_template('presupuesto_builder.html', page='presupuestos',
                           presupuesto=p, trabajadores=trabajadores)

@app.route('/api/presupuestos/add', methods=['POST'])
@login_required
def api_presupuestos_add():
    data = request.get_json()
    data['creado_por'] = session.get('usuario', '')
    ok, err = db.add_presupuesto(data)
    return jsonify({'ok': ok, 'id': err if ok else None, 'error': None if ok else err})

@app.route('/api/presupuestos/<doc_id>/edit', methods=['POST'])
@login_required
def api_presupuestos_edit(doc_id):
    data = request.get_json()
    data['editado_por'] = session.get('usuario', '')
    ok, err = db.update_presupuesto(doc_id, data)
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/presupuestos/<doc_id>/delete', methods=['POST'])
@login_required
def api_presupuestos_delete(doc_id):
    ok = db.delete_presupuesto(doc_id)
    return jsonify({'ok': ok})

@app.route('/api/presupuestos/<doc_id>/duplicar', methods=['POST'])
@login_required
def api_presupuestos_duplicar(doc_id):
    original = db.get_presupuesto_by_id(doc_id)
    if not original:
        return jsonify({'ok': False, 'error': 'Presupuesto no encontrado'})
    copia = {k: v for k, v in original.items() if k != 'id'}
    copia['estado'] = 'borrador'
    copia['creado_por'] = session.get('usuario', '')
    copia.pop('folio', None)
    ok, new_id = db.add_presupuesto(copia)
    return jsonify({'ok': ok, 'id': new_id if ok else None, 'error': None if ok else new_id})

@app.route('/admin/presupuestos/<doc_id>/pdf')
@login_required
def presupuesto_pdf(doc_id):
    p = db.get_presupuesto_by_id(doc_id)
    if not p:
        return '<h2>Presupuesto no encontrado</h2>', 404
    logo_b64 = get_logo_b64()
    sd = p.get('solar_data') or {}
    solar_kpis = _calcular_solar_kpis(sd, p.get('total', 0)) if sd.get('potencia_kwp') else {}
    equipos = _categorizar_items(p.get('items') or [])
    resp = make_response(render_template('presupuesto_pdf.html',
                                         p=p, logo_b64=logo_b64,
                                         solar_kpis=solar_kpis, equipos=equipos))
    if request.args.get('download'):
        nombre = (p.get('nombre_cliente','') or 'cliente').replace(' ', '_')
        folio  = p.get('folio', doc_id[:8])
        resp.headers['Content-Disposition'] = f'attachment; filename="Cotizacion_{folio}_{nombre}.html"'
        resp.headers['Content-Type'] = 'text/html'
    return resp

@app.route('/api/presupuestos/<doc_id>/enviar-email', methods=['POST'])
@login_required
def api_presupuestos_enviar_email(doc_id):
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    p = db.get_presupuesto_by_id(doc_id)
    if not p:
        return jsonify({'ok': False, 'error': 'Presupuesto no encontrado'})

    email_to = p.get('email_cliente', '').strip()
    if not email_to:
        return jsonify({'ok': False, 'error': 'El cliente no tiene email registrado'})

    if not EMAIL_PASS:
        return jsonify({'ok': False, 'error': 'EMAIL_PASS no configurada en las variables de entorno'})

    logo_b64 = get_logo_b64()
    sd_email = p.get('solar_data') or {}
    solar_kpis_email = _calcular_solar_kpis(sd_email, p.get('total', 0)) if sd_email.get('potencia_kwp') else {}
    equipos_email = _categorizar_items(p.get('items') or [])
    html_body = render_template('presupuesto_pdf.html', p=p, logo_b64=logo_b64,
                                 solar_kpis=solar_kpis_email, equipos=equipos_email)

    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"Presupuesto {p.get('folio','---')} — Corze Energía Solar"
        msg['From']    = f'Corze <{EMAIL_USER}>'
        msg['To']      = email_to
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))

        with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT) as server:
            server.starttls()
            server.login(EMAIL_USER, EMAIL_PASS)
            server.sendmail(EMAIL_USER, [email_to], msg.as_string())

        db.update_presupuesto(doc_id, {
            'email_enviado': True,
            'fecha_envio': db._now(),
            'estado': 'enviado',
        })
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# ══════════════════════════════════════════════════════════════════════════════
#  SOLAR — BOLETA + COTIZADOR
# ══════════════════════════════════════════════════════════════════════════════

_HSP = {
    'Arica y Parinacota':6.5,'Tarapacá':6.3,'Antofagasta':6.2,'Atacama':6.1,
    'Coquimbo':5.6,'Valparaíso':5.1,'Metropolitana de Santiago':5.0,
    "O'Higgins":4.8,'Maule':4.6,'Ñuble':4.4,'Biobío':4.2,
    'La Araucanía':3.9,'Los Ríos':3.6,'Los Lagos':3.3,'Aysén':2.9,'Magallanes':2.5,
}
_FACTOR_ESTACIONAL = {
    'Ene':1.2857,'Feb':1.2078,'Mar':1.1006,'Abr':0.9253,'May':0.7597,'Jun':0.6331,
    'Jul':0.6623,'Ago':0.7987,'Sep':0.9545,'Oct':1.1104,'Nov':1.2468,'Dic':1.3149,
}
_FACTOR_ORIENT = {
    'Norte':1.0,'Nororiente':0.95,'Norponiente':0.95,
    'Oriente':0.85,'Poniente':0.85,'Suroriente':0.5,'Surponiente':0.5,
    'Sur':0.4,'Horizontal':0.9,
}
_COMUNAS_REGION = {
    'LAS CONDES':'Metropolitana de Santiago','PROVIDENCIA':'Metropolitana de Santiago',
    'SANTIAGO':'Metropolitana de Santiago','VITACURA':'Metropolitana de Santiago',
    'ÑUÑOA':'Metropolitana de Santiago','LA REINA':'Metropolitana de Santiago',
    'MAIPÚ':'Metropolitana de Santiago','PUENTE ALTO':'Metropolitana de Santiago',
    'PEÑALOLÉN':'Metropolitana de Santiago','LA FLORIDA':'Metropolitana de Santiago',
    'SAN MIGUEL':'Metropolitana de Santiago','MACUL':'Metropolitana de Santiago',
    'HUECHURABA':'Metropolitana de Santiago','RECOLETA':'Metropolitana de Santiago',
    'COLINA':'Metropolitana de Santiago','BUIN':'Metropolitana de Santiago',
    'VALPARAÍSO':'Valparaíso','VIÑA DEL MAR':'Valparaíso','QUILPUÉ':'Valparaíso',
    'SAN ANTONIO':'Valparaíso','QUILLOTA':'Valparaíso',
    'CONCEPCIÓN':'Biobío','TALCAHUANO':'Biobío','CHILLÁN':'Ñuble',
    'TEMUCO':'La Araucanía','OSORNO':'Los Lagos','PUERTO MONTT':'Los Lagos',
    'IQUIQUE':'Tarapacá','ANTOFAGASTA':'Antofagasta','CALAMA':'Antofagasta',
    'LA SERENA':'Coquimbo','COQUIMBO':'Coquimbo','RANCAGUA':"O'Higgins",
    'TALCA':'Maule','CURICÓ':'Maule','VALDIVIA':'Los Ríos','ARICA':'Arica y Parinacota',
}

def _categorizar_items(items: list) -> dict:
    """Separa ítems en paneles, inversores, baterías y otros."""
    import re as _re
    paneles, inversores, baterias, otros = [], [], [], []
    for it in items:
        n = (it.get('nombre') or '').lower()
        if any(k in n for k in ['panel','bifacial','monocristalino','policristalino',
                                  'jinko','longi','canadiansolar','trina','yingli']) \
                or _re.search(r'\b\d{3,4}\s*w\b', n):
            # Extraer potencia unitaria para mostrar en propuesta
            m = _re.search(r'(\d{3,4})\s*[Ww]', it.get('nombre',''))
            it = dict(it)
            it['_spec'] = f"{m.group(1)} W c/u" if m else ''
            paneles.append(it)
        elif any(k in n for k in ['inversor','inverter','solis','growatt','huawei',
                                   'fronius','solaredge','deye','goodwe','[inv']):
            kw_m = _re.search(r'(\d+(?:[.,]\d+)?)\s*k[Ww]', it.get('nombre',''))
            it = dict(it)
            it['_spec'] = f"{kw_m.group(1)} kW" if kw_m else ''
            inversores.append(it)
        elif any(k in n for k in ['bater','battery','pylontech','byd','dyness','[bat']):
            baterias.append(it)
        else:
            otros.append(it)
    return {'paneles': paneles, 'inversores': inversores, 'baterias': baterias, 'otros': otros}


def _calcular_solar_kpis(sd: dict, total_proyecto: float = 0) -> dict:
    """Recalcula KPIs solares a partir del solar_data guardado en un presupuesto."""
    potencia_kwp = float(sd.get('potencia_kwp', 0))
    if potencia_kwp <= 0:
        return {}
    region       = sd.get('region', 'Metropolitana de Santiago')
    orientacion  = sd.get('orientacion', 'Norte')
    pr           = float(sd.get('pr', 80)) / 100
    tarifa_kwh   = float(sd.get('tarifa_kwh', 230))
    consumos     = sd.get('consumos', {})
    incremento   = 0.05
    degradacion  = 0.005

    hsp           = _HSP.get(region, 5.0)
    factor_orient = _FACTOR_ORIENT.get(orientacion, 1.0)
    MESES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']

    gen_anual = potencia_kwp * hsp * 30.4 * 12 * pr * factor_orient
    sum_f     = sum(_FACTOR_ESTACIONAL.values())

    gen_por_mes   = []
    consumo_anual = 0
    for mes in MESES:
        factor  = _FACTOR_ESTACIONAL.get(mes, 1.0)
        gen_mes = gen_anual * factor / sum_f
        c_mes   = float(consumos.get(mes, 0))
        consumo_anual += c_mes
        gen_por_mes.append({'mes': mes, 'consumo': round(c_mes, 1), 'generacion': round(gen_mes, 1)})

    consumo_mensual_prom = consumo_anual / 12 if consumo_anual > 0 else 0
    cobertura_pct        = gen_anual / consumo_anual * 100 if consumo_anual > 0 else 0

    proyeccion, ahorro_acum, neto_acum, payback = [], 0, -total_proyecto, None
    for anio in range(1, 26):
        gen_a = gen_anual * ((1 - degradacion) ** (anio - 1))
        tar_a = tarifa_kwh * ((1 + incremento) ** (anio - 1))
        ah_a  = gen_a * tar_a
        ahorro_acum += ah_a
        neto_acum   += ah_a
        if payback is None and neto_acum >= 0 and proyeccion:
            frac    = -(neto_acum - ah_a) / ah_a
            payback = anio - 1 + frac
        proyeccion.append({'anio': anio, 'gen': round(gen_a, 1),
                           'tarifa': round(tar_a, 1), 'ahorro': round(ah_a, 0),
                           'acumulado': round(ahorro_acum, 0), 'neto': round(neto_acum, 0)})

    co2_ton = gen_anual * 0.35 / 1000
    return {
        'gen_anual_kwh':        round(gen_anual, 1),
        'gen_mensual_kwh':      round(gen_anual / 12, 1),
        'consumo_anual':        round(consumo_anual, 1),
        'consumo_mensual_prom': round(consumo_mensual_prom, 1),
        'cobertura_pct':        round(cobertura_pct, 1),
        'payback_anios':        round(payback, 1) if payback is not None else None,
        'ahorro_anio1':         round(proyeccion[0]['ahorro'], 0) if proyeccion else 0,
        'ahorro_10':            round(proyeccion[9]['acumulado'], 0)  if len(proyeccion) >= 10 else 0,
        'ahorro_15':            round(proyeccion[14]['acumulado'], 0) if len(proyeccion) >= 15 else 0,
        'ahorro_25':            round(proyeccion[24]['acumulado'], 0) if len(proyeccion) >= 25 else 0,
        'co2_anual_ton':        round(co2_ton, 2),
        'hsp':                  hsp,
        'gen_por_mes':          gen_por_mes,
        'proyeccion':           proyeccion,
    }


@app.route('/api/boleta/leer', methods=['POST'])
@login_required
def api_leer_boleta():
    import re, io
    if 'archivo' not in request.files:
        return jsonify({'ok': False, 'error': 'No se recibió archivo'})
    archivo = request.files['archivo']
    if not archivo.filename.lower().endswith('.pdf'):
        return jsonify({'ok': False, 'error': 'El archivo debe ser PDF'})
    try:
        from pdfminer.high_level import extract_text
        texto = extract_text(io.BytesIO(archivo.read()))
        lineas = [l.strip() for l in texto.split('\n')
                  if l.strip() and not l.strip().startswith('(cid:')]
        txt = '\n'.join(lineas)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Error leyendo PDF: {e}'})

    datos = {
        'nombre':'','apellido':'','rut':'','direccion':'','comuna':'',
        'region':'','distribuidor':'','tarifa_tipo':'',
        'consumo_actual_kwh':0,'consumos_mensuales':{},'tarifa_kwh':230,'total_boleta':0,
    }

    # Nombre cliente
    m = re.search(r'Sr\.\s*\(a\)\s+([A-ZÁÉÍÓÚÜÑ][A-ZÁÉÍÓÚÜÑ\s]+?)[\s]*\.?\s*\n', txt, re.I)
    if m:
        full = m.group(1).strip().title()
        p = full.split()
        datos['nombre']   = p[0] if p else ''
        datos['apellido'] = ' '.join(p[1:]) if len(p)>1 else ''

    # Dirección
    m = re.search(r'Direcci[oó]n suministro:\s*(.+)', txt, re.I)
    if m:
        addr = m.group(1).strip()
        datos['direccion'] = addr.title()
        parts = addr.upper().rsplit('-', 1)
        if len(parts) == 2:
            comuna = parts[1].strip()
            datos['comuna'] = comuna.title()
            datos['region'] = _COMUNAS_REGION.get(comuna, '')

    # RUT (del cliente si aparece explícito)
    m = re.search(r'R\.?U\.?T\.?\s*cliente[:\s]+(\d[\d\.]+\-[\dkK])', txt, re.I)
    if m: datos['rut'] = m.group(1)

    # Tarifa tipo
    m = re.search(r'Tipo de tarifa contratada:\s*(\S+)', txt, re.I)
    if m: datos['tarifa_tipo'] = m.group(1)

    # Consumo actual del periodo
    m = re.search(r'Electricidad Consumida\s*\((\d+)\s*kWh\)', txt, re.I)
    if not m:
        m = re.search(r'Consumo total del periodo\s*=\s*(\d+)\s*kWh', txt, re.I)
    if m: datos['consumo_actual_kwh'] = int(m.group(1))

    # Total a pagar
    m = re.search(r'Total a pagar[:\s]*\$\s*([\d\.]+)', txt, re.I)
    if m: datos['total_boleta'] = int(m.group(1).replace('.',''))

    # Costo electricidad (para calcular tarifa)
    m = re.search(r'Electricidad Consumida[^\n]*\n[^\$]*\$\s*([\d\.]+)', txt, re.I)
    if m and datos['consumo_actual_kwh']:
        costo_elec = int(m.group(1).replace('.',''))
        datos['tarifa_kwh'] = round(costo_elec / datos['consumo_actual_kwh'])

    # Distribuidora
    for dist in ['Enel','CGE','Chilquinta','Frontel','Saesa','Edelaysen','Edelmag']:
        if dist.lower() in txt.lower():
            datos['distribuidor'] = dist; break

    # ── Consumos mensuales del historial ─────────────────────────────────────
    # Strategy:
    #   1. Find the X-axis months line (≥5 month abbrevs on one line)
    #   2. Extract numbers ONLY from the chart section (between "Consumo total
    #      del periodo" and the months line) → avoids picking up year "2026",
    #      address numbers, billing amounts from the rest of the document
    #   3. Strip Y-axis scale values (always multiples of 50: 200,400,600,800,1000…)
    #   4. Map bar values sequentially to months
    #   5. Override current month and previous month with reliable anchor values
    #      extracted from the comparison mini-charts after the months line
    _MESES_KEY  = {'ENE':'Ene','FEB':'Feb','MAR':'Mar','ABR':'Abr','MAY':'May','JUN':'Jun',
                   'JUL':'Jul','AGO':'Ago','SEP':'Sep','OCT':'Oct','NOV':'Nov','DIC':'Dic'}
    _MESES_NORM = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    _MESES_ABR  = ['ENE','FEB','MAR','ABR','MAY','JUN','JUL','AGO','SEP','OCT','NOV','DIC']

    pat_mln = (r'((?:(?:ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)\s+){4,}'
               r'(?:ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC))')
    m_ml = re.search(pat_mln, txt, re.I)

    if m_ml:
        meses_linea = m_ml.group(1).strip().upper().split()
        pos_ml_start = m_ml.start()
        pos_ml_end   = m_ml.end()

        # Chart section: from right after "Consumo total del periodo…\n" to month line
        m_ctdp = re.search(r'Consumo\s+total\s+del\s+periodo[^\n]*\n', txt[:pos_ml_start], re.I)
        sec_start = m_ctdp.end() if m_ctdp else max(0, pos_ml_start - 800)
        chart_sec = txt[sec_start:pos_ml_start]

        # Raw numbers 50-2500 from chart section only
        chart_raw = [int(v) for v in re.findall(r'\b(\d{2,4})\b', chart_sec)
                     if 50 <= int(v) <= 2500]

        # Y-axis values in Enel/CGE charts are ALWAYS multiples of 50 (200,400,600…)
        # Skip the leading Y-axis block; keep values not divisible by 50
        bar_vals = [v for v in chart_raw if v % 50 != 0]

        # Sequential month → value mapping
        consumos = {}
        for i, mes_k in enumerate(meses_linea):
            if i < len(bar_vals):
                consumos[_MESES_KEY.get(mes_k, mes_k.title())] = bar_vals[i]

        # ── Anchor overrides from comparison mini-charts (after month labels) ──
        # Post-months text typically: current_val, same_yr_val, …, current_val, prev_val, …
        texto_post = txt[pos_ml_end:]
        post_raw = [int(v) for v in re.findall(r'\b(\d{3,4})\b', texto_post[:600])
                    if 50 <= int(v) <= 5000 and int(v) % 50 != 0]
        # post_raw should be: [current, same_yr_ago, current, last_month, …]

        m_fecha = re.search(r'Fecha de emisi[oó]n:\s*\d+\s+(\w+)\s+(\d{4})', txt, re.I)
        consumo_actual = datos.get('consumo_actual_kwh', 0)
        if m_fecha and consumo_actual:
            mes_raw = m_fecha.group(1)[:3].upper()
            if mes_raw in _MESES_ABR:
                idx_act   = _MESES_ABR.index(mes_raw)
                mes_act_k = _MESES_NORM[idx_act]
                mes_ant_k = _MESES_NORM[(idx_act - 1) % 12]
                # Override current month with the known-correct bill value
                consumos[mes_act_k] = consumo_actual
                # Override previous month from comparison mini-chart
                if len(post_raw) >= 4:
                    consumos[mes_ant_k] = post_raw[3]

        datos['consumos_mensuales'] = consumos

    datos['ok'] = True
    return jsonify(datos)


@app.route('/api/cotizacion/calcular', methods=['POST'])
@login_required
def api_calcular_cotizacion():
    d = request.get_json() or {}
    potencia_kwp   = float(d.get('potencia_kwp', 0))
    region         = d.get('region', 'Metropolitana de Santiago')
    orientacion    = d.get('orientacion', 'Norte')
    pr             = float(d.get('pr', 80)) / 100
    tarifa_kwh     = float(d.get('tarifa_kwh', 230))
    precio_excedente = float(d.get('precio_excedente', 110))
    incremento     = float(d.get('incremento_anual', 5)) / 100
    degradacion    = float(d.get('degradacion_anual', 0.5)) / 100
    total_proyecto = float(d.get('total_proyecto', 0))
    consumos       = d.get('consumos', {})  # {Ene:947, Feb:829, ...}

    if potencia_kwp <= 0:
        return jsonify({'ok': False, 'error': 'Potencia del sistema requerida'})

    hsp           = _HSP.get(region, 5.0)
    factor_orient = _FACTOR_ORIENT.get(orientacion, 1.0)
    MESES         = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']

    # Generación anual: P_kWp × HSP × 30.4 días × 12 meses × PR × factor_orient
    gen_anual = potencia_kwp * hsp * 30.4 * 12 * pr * factor_orient
    sum_f     = sum(_FACTOR_ESTACIONAL.values())  # ≈ 12

    gen_por_mes = []
    consumo_anual = 0
    for mes in MESES:
        factor  = _FACTOR_ESTACIONAL.get(mes, 1.0)
        gen_mes = gen_anual * factor / sum_f
        c_mes   = float(consumos.get(mes, 0))
        consumo_anual += c_mes
        gen_por_mes.append({'mes': mes, 'consumo': round(c_mes, 1), 'generacion': round(gen_mes, 1)})

    consumo_mensual_prom = consumo_anual / 12 if consumo_anual > 0 else 0
    cobertura_pct        = gen_anual / consumo_anual * 100 if consumo_anual > 0 else 0

    # Proyección 25 años
    proyeccion, ahorro_acum, neto_acum, payback = [], 0, -total_proyecto, None
    for anio in range(1, 26):
        gen_a  = gen_anual * ((1 - degradacion) ** (anio - 1))
        tar_a  = tarifa_kwh * ((1 + incremento) ** (anio - 1))
        ah_a   = gen_a * tar_a
        ahorro_acum += ah_a
        neto_acum   += ah_a
        if payback is None and neto_acum >= 0 and proyeccion:
            frac   = -(neto_acum - ah_a) / ah_a
            payback = anio - 1 + frac
        proyeccion.append({'anio': anio, 'gen': round(gen_a,1),
                           'tarifa': round(tar_a,1), 'ahorro': round(ah_a,0),
                           'acumulado': round(ahorro_acum,0), 'neto': round(neto_acum,0)})

    co2_ton = gen_anual * 0.35 / 1000

    return jsonify({
        'ok': True,
        'gen_mensual_kwh':     round(gen_anual / 12, 1),
        'gen_anual_kwh':       round(gen_anual, 1),
        'consumo_mensual_prom':round(consumo_mensual_prom, 1),
        'consumo_anual':       round(consumo_anual, 1),
        'cobertura_pct':       round(cobertura_pct, 1),
        'payback_anios':       round(payback, 1) if payback is not None else None,
        'ahorro_anio1':        round(proyeccion[0]['ahorro'], 0) if proyeccion else 0,
        'ahorro_10':           round(proyeccion[9]['acumulado'], 0) if len(proyeccion) >= 10 else 0,
        'ahorro_25':           round(proyeccion[24]['acumulado'], 0) if len(proyeccion) >= 25 else 0,
        'roi_25':              round(proyeccion[24]['acumulado'] / total_proyecto * 100, 1) if total_proyecto > 0 and len(proyeccion) >= 25 else 0,
        'co2_anual_ton':       round(co2_ton, 2),
        'hsp':                 hsp,
        'gen_por_mes':         gen_por_mes,
        'proyeccion':          proyeccion,
    })

# ══════════════════════════════════════════════════════════════════════════════
#  CRM PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

ETAPAS_PIPELINE = [
    'Nuevo Lead',
    'Intento Llamado 1°',
    'Intento Llamado 2°',
    'Contactado',
    'Esperando Cuenta Luz',
    'Visita Técnica Agendada',
    'Propuesta Enviada',
    'En Negociación',
    'Esperando Documentación',
    'Contrato Firmado',
    'En Instalación',
    'Proyecto Finalizado',
    'Post Venta',
    'Cliente Perdido',
]

@app.route('/admin/pipeline')
@login_required
def pipeline():
    leads = db.get_all_leads()
    trabajadores = db.get_all_trabajadores(solo_activos=True)
    return render_template('pipeline.html', page='pipeline',
                           leads=leads, etapas=ETAPAS_PIPELINE,
                           trabajadores=trabajadores)

@app.route('/api/leads/add', methods=['POST'])
@login_required
def api_leads_add():
    data = request.get_json()
    data['creado_por'] = session.get('usuario', '')
    ok, err = db.add_lead(data)
    return jsonify({'ok': ok, 'id': err if ok else None, 'error': None if ok else err})

@app.route('/api/leads', methods=['GET'])
@login_required
def api_leads_list():
    etapa    = request.args.get('etapa', '')
    asignado = request.args.get('asignado', '')
    q        = request.args.get('q', '')
    rows = db.get_all_leads(etapa=etapa, asignado=asignado, query=q)
    return jsonify(rows)

@app.route('/api/leads/<doc_id>', methods=['GET'])
@login_required
def api_leads_get(doc_id):
    return jsonify(db.get_lead_by_id(doc_id) or {})

@app.route('/api/leads/<doc_id>/edit', methods=['POST'])
@login_required
def api_leads_edit(doc_id):
    data = request.get_json()
    data['editado_por'] = session.get('usuario', '')
    ok, err = db.update_lead(doc_id, data)
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/leads/<doc_id>/delete', methods=['POST'])
@login_required
def api_leads_delete(doc_id):
    ok = db.delete_lead(doc_id)
    return jsonify({'ok': ok})

@app.route('/api/leads/<doc_id>/mover-etapa', methods=['POST'])
@login_required
def api_leads_mover_etapa(doc_id):
    data      = request.get_json()
    nueva_etapa = data.get('etapa', '')
    if nueva_etapa not in ETAPAS_PIPELINE:
        return jsonify({'ok': False, 'error': 'Etapa inválida'})
    lead = db.get_lead_by_id(doc_id)
    if not lead:
        return jsonify({'ok': False, 'error': 'Lead no encontrado'})
    etapa_anterior = lead.get('etapa', '')
    ok, err = db.update_lead(doc_id, {'etapa': nueva_etapa})
    if ok:
        db.add_historial_lead(doc_id, 'Cambio de etapa',
                              session.get('usuario', ''),
                              f'{etapa_anterior} → {nueva_etapa}')
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/leads/<doc_id>/notas', methods=['POST'])
@login_required
def api_leads_add_nota(doc_id):
    data  = request.get_json()
    texto = (data.get('texto') or '').strip()
    if not texto:
        return jsonify({'ok': False, 'error': 'Texto requerido'})
    nota = {
        'texto':  texto,
        'autor':  session.get('usuario', 'Sistema'),
        'fecha':  _cl().strftime('%Y-%m-%d %H:%M'),
    }
    ok = db.add_nota_lead(doc_id, nota)
    return jsonify({'ok': ok, 'nota': nota})

@app.route('/api/leads/<doc_id>/presupuestos', methods=['GET'])
@login_required
def api_leads_presupuestos(doc_id):
    """Devuelve los presupuestos asociados a un lead (por email o lead_id)."""
    lead = db.get_lead_by_id(doc_id)
    if not lead:
        return jsonify([])
    email = (lead.get('email') or '').strip().lower()
    todos = db.get_all_presupuestos()
    match = []
    for p in todos:
        p_email = (p.get('email_cliente') or '').strip().lower()
        if (email and p_email and email == p_email) or p.get('lead_id') == doc_id:
            match.append({'id': p['id'], 'folio': p.get('folio',''), 'estado': p.get('estado',''),
                          'total': p.get('total',0), 'created_at': p.get('created_at','')})
    return jsonify(match)

@app.route('/api/leads/<doc_id>/grabaciones', methods=['GET'])
@login_required
def api_lead_grabaciones_list(doc_id):
    return jsonify(db.get_grabaciones_lead(doc_id))

@app.route('/api/leads/<doc_id>/grabaciones', methods=['POST'])
@login_required
def api_lead_grabaciones_add(doc_id):
    data = request.get_json(force=True)
    audio_b64 = data.get('audio_b64', '')
    # Validar tamaño (base64: cada 4 chars = 3 bytes raw)
    raw_kb = len(audio_b64) * 3 // 4 // 1024
    if raw_kb > 800:
        return jsonify({'ok': False, 'error': f'Grabación demasiado larga ({raw_kb} KB). Máximo ~800 KB.'}), 400
    payload = {
        'duracion':       data.get('duracion', 0),
        'nombre_archivo': data.get('nombre_archivo', 'grabacion.webm'),
        'usuario':        data.get('usuario', session.get('usuario', '')),
        'notas':          data.get('notas', ''),
        'audio_b64':      audio_b64,
    }
    ok, rec_id = db.add_grabacion_lead(doc_id, payload)
    return jsonify({'ok': ok, 'id': rec_id if ok else None, 'error': rec_id if not ok else None})

@app.route('/api/leads/<doc_id>/grabaciones/<rec_id>', methods=['GET'])
@login_required
def api_lead_grabacion_get(doc_id, rec_id):
    rec = db.get_grabacion(rec_id)
    return jsonify(rec)

@app.route('/api/leads/<doc_id>/grabaciones/<rec_id>', methods=['DELETE'])
@login_required
def api_lead_grabacion_delete(doc_id, rec_id):
    ok = db.delete_grabacion(rec_id)
    return jsonify({'ok': ok})

@app.route('/api/leads/importar-email', methods=['POST'])
@login_required
def api_leads_importar_email():
    """Parsea texto de email y crea un lead."""
    data = request.get_json()
    texto = data.get('texto', '')
    if not texto:
        return jsonify({'ok': False, 'error': 'Texto vacío'})

    import re
    lead_data = {
        'nombre': '',
        'apellido': '',
        'email': '',
        'telefono': '',
        'empresa': '',
        'origen': 'email',
        'notas': texto[:500],
        'etapa': 'Nuevo Lead',
        'creado_por': session.get('usuario', ''),
    }

    email_match = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', texto)
    if email_match:
        lead_data['email'] = email_match.group(0)

    phone_match = re.search(r'(\+?56\s?9\s?\d{4}\s?\d{4}|9\d{8}|\d{9})', texto)
    if phone_match:
        lead_data['telefono'] = phone_match.group(0).replace(' ', '')

    name_match = re.search(r'(?:nombre[:\s]+|de[:\s]+|from[:\s]+)([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?:\s+[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+)*)', texto, re.IGNORECASE)
    if name_match:
        parts = name_match.group(1).strip().split()
        lead_data['nombre'] = parts[0] if parts else ''
        lead_data['apellido'] = ' '.join(parts[1:]) if len(parts) > 1 else ''
    else:
        lead_data['nombre'] = 'Lead Email'

    ok, err = db.add_lead(lead_data)
    return jsonify({'ok': ok, 'id': err if ok else None, 'error': None if ok else err})


@app.route('/api/leads/importar-eml', methods=['POST'])
@login_required
def api_leads_importar_eml():
    """Parsea un archivo .eml de formulario SolarLaunch y crea un lead."""
    import re as _re
    import email as _eml

    try:
        # ── 1. Validar archivo ─────────────────────────────────────────────
        if 'archivo' not in request.files:
            return jsonify({'ok': False, 'error': 'No se recibió archivo (campo: archivo)'})
        f = request.files['archivo']
        if not f.filename.lower().endswith('.eml'):
            return jsonify({'ok': False, 'error': 'El archivo debe ser .eml'})

        raw = f.read()
        if not raw:
            return jsonify({'ok': False, 'error': 'El archivo está vacío'})

        # ── 2. Parsear MIME ────────────────────────────────────────────────
        msg = _eml.message_from_bytes(raw)

        asunto = msg.get('Subject', '') or ''
        try:
            from email.header import decode_header as _dh
            asunto = ' '.join(
                (p.decode(enc or 'utf-8') if isinstance(p, bytes) else p)
                for p, enc in _dh(asunto)
            )
        except Exception:
            pass

        # ── 3. Extraer body (plain text preferido, fallback HTML) ──────────
        body = ''
        for part in msg.walk():
            if part.get_content_type() == 'text/plain':
                payload = part.get_payload(decode=True)
                if payload:
                    body = payload.decode(part.get_content_charset() or 'utf-8', errors='replace')
                    break
        if not body:
            for part in msg.walk():
                if part.get_content_type() == 'text/html':
                    payload = part.get_payload(decode=True)
                    if payload:
                        html = payload.decode(part.get_content_charset() or 'utf-8', errors='replace')
                        body = _re.sub(r'<br\s*/?>', '\n', html, flags=_re.IGNORECASE)
                        body = _re.sub(r'<[^>]+>', '', body)
                        break

        if not body:
            return jsonify({'ok': False, 'error': 'No se encontró texto en el EML'})

        # ── 4. Normalizar saltos de línea ──────────────────────────────────
        body = body.replace('\r\n', '\n').replace('\r', '\n')

        # ── 5. Parsear pares pregunta / respuesta ──────────────────────────
        nombre = apellido = email_lead = telefono = ciudad = region = ''
        consumo_estimado = tipo_proyecto = ''

        for block in body.split('\n\n'):
            block = block.strip()
            if not block:
                continue
            lines = [l.strip() for l in block.split('\n') if l.strip()]
            if len(lines) < 2:
                continue
            q = lines[0].lower()
            a = lines[1]
            if not a:
                continue

            if any(k in q for k in ('nombre', 'name')):
                parts = a.split()
                nombre   = parts[0] if parts else a
                apellido = ' '.join(parts[1:]) if len(parts) > 1 else ''
            elif any(k in q for k in ('email', 'correo', 'e-mail')):
                email_lead = a.strip()
            elif any(k in q for k in ('tel', 'whatsapp', 'phone', 'fono', 'celular')):
                telefono = _re.sub(r'\s+', '', a)
            elif any(k in q for k in ('ciudad', 'indica', 'localidad', 'city')):
                ciudad = a
            elif any(k in q for k in ('regi', 'region', 'región')):
                region = a
            elif any(k in q for k in ('pagas', 'electricidad', 'boleta', 'luz', 'gasto')):
                consumo_estimado = a
            elif any(k in q for k in ('donde', 'dónde', 'evaluar', 'soluci', 'instala', 'proyecto', 'residencial', 'comercial')):
                tipo_proyecto = ('Residencial'
                                 if any(w in a.lower() for w in ('casa', 'residencial', 'hogar', 'departamento'))
                                 else 'Comercial')

        # ── 6. Fallback: extraer email por regex ───────────────────────────
        if not email_lead:
            m = _re.search(r'[\w\.\-]+@[\w\.\-]+\.\w+', body)
            if m:
                email_lead = m.group(0)

        if not nombre:
            nombre = 'Lead Web'

        # ── 7. Verificar duplicado ─────────────────────────────────────────
        if email_lead:
            existing = db.get_lead_by_email(email_lead)
            if existing:
                db.add_historial_lead(
                    existing['id'], 'EML duplicado',
                    session.get('usuario', ''),
                    f'EML subido pero {email_lead} ya existe.'
                )
                return jsonify({
                    'ok': True, 'duplicado': True,
                    'msg': f"Ya existe: {existing.get('nombre','')} {existing.get('apellido','')}".strip(),
                    'lead_id': existing['id'],
                    'nombre': existing.get('nombre', ''),
                    'email': email_lead,
                })

        # ── 8. Crear lead ──────────────────────────────────────────────────
        notas_parts = []
        if tipo_proyecto:    notas_parts.append(f'Tipo proyecto: {tipo_proyecto}')
        if consumo_estimado: notas_parts.append(f'Consumo estimado: {consumo_estimado}')
        if ciudad:           notas_parts.append(f'Ciudad: {ciudad}')
        if region:           notas_parts.append(f'Región: {region}')
        if asunto:           notas_parts.append(f'Asunto EML: {asunto}')

        lead_data = {
            'nombre':           nombre,
            'apellido':         apellido,
            'email':            email_lead,
            'telefono':         telefono,
            'ciudad':           ciudad,
            'region':           region,
            'tipo_proyecto':    tipo_proyecto,
            'consumo_estimado': consumo_estimado,
            'empresa':          '',
            'origen':           'formulario_web',
            'notas':            '\n'.join(notas_parts),
            'etapa':            'Nuevo Lead',
            'creado_por':       session.get('usuario', ''),
        }
        ok, err = db.add_lead(lead_data)
        if not ok:
            return jsonify({'ok': False, 'error': f'Error Firestore: {err}'})

        return jsonify({
            'ok': True,
            'id': err,
            'nombre': nombre,
            'apellido': apellido,
            'email': email_lead,
            'telefono': telefono,
            'tipo_proyecto': tipo_proyecto,
            'consumo_estimado': consumo_estimado,
        })

    except Exception as exc:
        import traceback
        return jsonify({'ok': False, 'error': f'Error interno: {exc}', 'trace': traceback.format_exc()[-400:]})


# ══════════════════════════════════════════════════════════════════════════════
#  TRABAJADORES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/admin/trabajadores')
@login_required
def trabajadores():
    rows = db.get_all_trabajadores(solo_activos=False)
    return render_template('trabajadores.html', page='trabajadores', trabajadores=rows)

@app.route('/api/trabajadores', methods=['GET'])
@login_required
def api_trabajadores_list():
    solo_activos = request.args.get('solo_activos', 'true').lower() != 'false'
    rows = db.get_all_trabajadores(solo_activos=solo_activos)
    return jsonify(rows)

@app.route('/api/trabajadores/add', methods=['POST'])
@login_required
def api_trabajadores_add():
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'ok': False, 'error': 'Datos vacíos o Content-Type incorrecto'})
        if not data.get('nombre', '').strip():
            return jsonify({'ok': False, 'error': 'El nombre es requerido'})
        if not data.get('avatar'):
            n = (data.get('nombre', '') + ' ' + data.get('apellido', '')).strip()
            parts = n.split()
            data['avatar'] = ''.join(p[0].upper() for p in parts[:2]) if parts else '??'
        ok, err = db.add_trabajador(data)
        return jsonify({'ok': ok, 'id': err if ok else None, 'error': None if ok else err})
    except Exception as exc:
        import traceback
        return jsonify({'ok': False, 'error': f'Error interno: {exc}', 'trace': traceback.format_exc()[-300:]})

@app.route('/api/trabajadores/<doc_id>/edit', methods=['POST'])
@login_required
def api_trabajadores_edit(doc_id):
    data = request.get_json()
    ok, err = db.update_trabajador(doc_id, data)
    return jsonify({'ok': ok, 'error': err})

@app.route('/api/trabajadores/<doc_id>/delete', methods=['POST'])
@login_required
def api_trabajadores_delete(doc_id):
    ok = db.delete_trabajador(doc_id)
    return jsonify({'ok': ok})
